"""Build the rights-neutral blank SBOM template used by the public source set."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PUBLIC_TEMPLATE_FILENAME = "客户SBOM导入模板_公开空白版.xlsx"
SBOM_HEADERS = [
    "Row ID 行号",
    "Component category 组件类别",
    "Component producer 组件生产者",
    "Component name 组件名称",
    "Component version 组件版本",
    "PURL Package URL",
    "CPE",
    "Internal ID 内部标识",
    "Dependency relationship 依赖关系",
    "Source / Evidence 来源/证据",
    "Used in product build 是否进入目标构建",
    "Security relevance 安全相关性",
    "Known uncertainty / gap 已知不确定性/缺口",
    "License 许可证",
    "CVE 漏洞编号",
    "EUVD ID EUVD编号",
    "Customer notes 客户备注",
]


def _style_region(sheet, min_row: int, max_row: int, max_col: int) -> None:
    thin = Side(style="thin", color="B7C9D6")
    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header(sheet, width: int) -> None:
    for cell in sheet[1][:width]:
        cell.fill = PatternFill("solid", fgColor="163B65")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _add_table(sheet, name: str, reference: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    sheet.add_table(table)


def build_template_workbook() -> Workbook:
    workbook = Workbook()
    metadata = workbook.active
    metadata.title = "01_Metadata_元数据"
    sbom = workbook.create_sheet("02_SBOM_Software")
    instructions = workbook.create_sheet("04_Instructions_说明")

    metadata_rows = [
        ["Field 字段", "Customer entry 客户填写", "Example 示例", "Notes 说明"],
        ["Template source 模板来源", "", "Public blank SBOM template", "Rights-neutral template generated from this repository's public source."],
        ["Product name 产品名称", "", "SGW-200", "Product name or model."],
        ["Product version 产品版本", "", "FW 3.2.0", "Delivered software or firmware version."],
        ["Hardware revision 硬件修订", "", "HW Rev B", "Hardware, board, or module revision."],
        ["Build ID 构建号", "", "build-sgw200-3.2.0", "Unique delivered-build identifier."],
        ["Release date 发布日期", "", "2030-01-01", "Release date for this inventory."],
        ["SBOM author 编制者", "", "Product configuration owner", "Prefer a role or organization over personal data."],
        ["Contact 联系方式", "", "security@example.test", "Contact for clarification."],
        ["Inventory scope 清单范围", "", "Release SBOM", "State the release/internal/request view."],
        ["SBOM version SBOM版本", "", "1", "SBOM document version, not product version."],
        ["SBOM timestamp SBOM时间戳", "", "ISO 8601", "Creation timestamp with timezone."],
    ]
    for row in metadata_rows:
        metadata.append(row)
    _style_region(metadata, 1, len(metadata_rows), 4)
    _style_header(metadata, 4)
    for column, width in zip("ABCD", (30, 34, 34, 72), strict=True):
        metadata.column_dimensions[column].width = width
    metadata.freeze_panes = "A2"
    _add_table(metadata, "SbomMetadataInput", f"A1:D{len(metadata_rows)}")

    sbom.append(SBOM_HEADERS)
    sbom.append([""] * len(SBOM_HEADERS))
    _style_region(sbom, 1, 20, len(SBOM_HEADERS))
    _style_header(sbom, len(SBOM_HEADERS))
    widths = (12, 24, 24, 24, 18, 40, 38, 24, 32, 30, 22, 24, 34, 18, 24, 24, 30)
    for index, width in enumerate(widths, start=1):
        sbom.column_dimensions[sbom.cell(1, index).column_letter].width = width
    sbom.freeze_panes = "A2"
    category = DataValidation(
        type="list",
        formula1='"System kernel 系统内核,RTOS / BSP 系统软件,Third-party component 第三方组件,Self-developed code 自研代码,Supplier package 供应商封装组件"',
    )
    build_use = DataValidation(type="list", formula1='"Yes 是,No 否,Unknown 待确认"')
    sbom.add_data_validation(category)
    sbom.add_data_validation(build_use)
    category.add("B2:B200")
    build_use.add("K2:K200")
    for row in range(2, 201):
        for column in (6, 7, 8, 15, 16):
            sbom.cell(row, column).number_format = "@"
    _add_table(sbom, "SbomSoftwareInput", "A1:Q2")

    instruction_rows = [
        ["Topic 主题", "Guidance 指引", "Why it matters 重要性"],
        ["Start here 开始填写", "填写元数据，并在 SBOM 页每个组件填写一行；模板有意保持空白。", "避免把演示数据误当成交付产品证据。"],
        ["Do not guess versions 不要猜测版本", "未知版本填写 UNKNOWN，并记录缺口、责任人与下一步。", "猜测会导致错误匹配和审核证据。"],
        ["Use product/build binding 使用产品/构建绑定", "将 SBOM 与产品版本、硬件修订、软件构建及发布包绑定。", "证明实际交付对象。"],
        ["Use stable identifiers 使用稳定标识", "可用时分别填写 PURL、CPE 或内部 ID。", "支持确定性映射并减少误匹配。"],
        ["CVE and EUVD", "仅填写已经确认的标识；命中是候选信号。", "命中不单独证明产品适用性或 CRA Art.14 可报告性。"],
        ["Dependency relationship 依赖关系", "说明直接/间接依赖及父组件或构建关系。", "支持影响分析与追溯。"],
        ["New build or release 新构建或发布", "依赖发生变化时建立新的 SBOM 版本。", "避免复用其他配置的证据。"],
        ["Assessment boundary 评估边界", "自动输出只支持证据收集与漏洞分诊；完整性、适用性、利用证据、CRA/SRP、符合性与发布仍需人工复核。", "成功导入或匹配不等于自动合规或报告决定。"],
    ]
    for row in instruction_rows:
        instructions.append(row)
    _style_region(instructions, 1, len(instruction_rows), 3)
    _style_header(instructions, 3)
    for column, width in zip("ABC", (34, 88, 72), strict=True):
        instructions.column_dimensions[column].width = width
    instructions.freeze_panes = "A2"
    _add_table(instructions, "SbomInstructions", f"A1:C{len(instruction_rows)}")
    return workbook


def template_bytes() -> bytes:
    output = io.BytesIO()
    build_template_workbook().save(output)
    return output.getvalue()


def write_public_template(output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(template_bytes())
