"""Tests for the Phase B compliance evidence package.

Phase B adds a per-job evidence bundle export (CRA compliance deliverable).
These tests start with the audit-chain layer (workflow_store.verify_audit_chain
+ export_audit_trail); builder/writer/endpoint tests are added alongside their
implementation.
"""

from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from app.evidence_package import (
    build_evidence_package_payload,
    write_evidence_package_zip,
    write_evidence_xlsx,
)
from app.workflow_store import WorkflowStore


class AuditChainExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.store = WorkflowStore(Path(self._tmp.name) / "wb.sqlite3")

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _audit(self, case_id, event_type, actor, payload):
        with self.store.connect() as conn:
            self.store._audit(conn, case_id, event_type, actor, payload)

    def test_verify_audit_chain_clean(self) -> None:
        self._audit(None, "evt", "alice", {"n": 1})
        self._audit(None, "evt", "bob", {"n": 2})
        result = self.store.verify_audit_chain()
        self.assertTrue(result["verified"])
        self.assertIsNone(result["broken_at"])
        self.assertEqual(result["event_count"], 2)

    def test_verify_audit_chain_detects_tamper(self) -> None:
        self._audit(None, "evt", "alice", {"n": 1})
        self._audit(None, "evt", "bob", {"n": 2})
        with self.store.connect() as conn:
            conn.execute(
                "UPDATE audit_events SET payload_json=? WHERE id=1",
                ('{"n":99}',),
            )
        result = self.store.verify_audit_chain()
        self.assertFalse(result["verified"])
        self.assertEqual(result["broken_at"], 1)

    def test_verify_audit_chain_empty_is_verified(self) -> None:
        result = self.store.verify_audit_chain()
        self.assertTrue(result["verified"])
        self.assertEqual(result["event_count"], 0)

    def test_export_audit_trail_id_asc_eight_columns(self) -> None:
        self._audit(None, "evt", "alice", {"n": 1})
        self._audit(None, "evt", "bob", {"n": 2})
        rows = self.store.export_audit_trail()
        self.assertEqual([r["id"] for r in rows], [1, 2])  # id ASC, not DESC
        for col in (
            "id",
            "case_id",
            "event_type",
            "actor",
            "payload_json",
            "payload_sha256",
            "prev_hash",
            "created_at",
        ):
            self.assertIn(col, rows[0])
        self.assertEqual(rows[0]["prev_hash"], "")
        self.assertEqual(rows[1]["prev_hash"], rows[0]["payload_sha256"])

    def test_export_audit_trail_scoped_by_job(self) -> None:
        case = self.store.create_manual_case(
            {
                "project_name": "Gateway",
                "project_version": "1.0",
                "software_build": "b1",
                "component_name": "lib",
                "component_version": "1.0",
                "cve_id": "CVE-2026-1",
                "euvd_id": "EUVD-2026-1",
            },
            "analyst",
        )
        with self.store.connect() as conn:
            conn.execute(
                "UPDATE cases SET job_id=? WHERE id=?", ("job-1", case["id"])
            )
        self._audit(case["id"], "awareness_manually_confirmed", "alice", {"k": 1})
        self._audit(None, "unrelated", "system", {"k": 2})  # no case → not in job-1
        scoped = self.store.export_audit_trail("job-1")
        # every scoped row belongs to the job's case; the unrelated event is excluded
        self.assertTrue(scoped)
        self.assertTrue(all(r["case_id"] == case["id"] for r in scoped))
        self.assertFalse(
            any(r["event_type"] == "unrelated" for r in scoped)
        )
        full = self.store.export_audit_trail()
        self.assertGreater(len(full), len(scoped))  # full includes the unrelated row


class EvidencePayloadBuilderTests(unittest.TestCase):
    """Phase B build_evidence_package_payload — pure dict-in/dict-out."""

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        self.uploads = self.root / "uploads"
        self.uploads.mkdir()
        self.store = WorkflowStore(self.root / "wb.sqlite3")

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def _write_upload(self, upload_id: str, content: bytes, source_sha: str) -> None:
        (self.uploads / f"{upload_id}.bin").write_bytes(content)
        (self.uploads / f"{upload_id}.upload-record.json").write_text(
            json.dumps({"stored_file": f"{upload_id}.bin", "source_sha256": source_sha}),
            encoding="utf-8",
        )

    def _job(self, matches=None, components=None, source_sha="sha", upload_id="up-1") -> dict:
        return {
            "id": "job-1",
            "upload_id": upload_id,
            "project_name": "Gateway",
            "project_version": "1.0",
            "customer": "CustomerA",
            "software_build": "b1",
            "source_sha256": source_sha,
            "status": "completed",
            "result": {
                "matches": matches or [],
                "components": components or [],
                "summary": {"component_count": 2, "confirmed_findings": 1},
                "data_provenance": {"snapshot_sha256": "snap-sha", "status": "local_ready"},
            },
        }

    def test_project_identity_and_summary(self) -> None:
        payload = build_evidence_package_payload(self._job(), self.store, self.uploads)
        self.assertEqual(payload["schema_version"], "1.0")
        self.assertEqual(payload["project_identity"]["project_name"], "Gateway")
        self.assertEqual(payload["summary"]["component_count"], 2)

    def test_sbom_source_declarations_flow_through_payload(self) -> None:
        job = self._job()
        job["sbom_source_declarations"] = {
            "classification": "SELF_TEST_NOT_CUSTOMER_EVIDENCE",
            "direction": "SBOM_TO_EUVD_ONLY",
            "source_binding_status": "CALLER_DECLARED_NOT_INDEPENDENTLY_VERIFIED",
        }
        payload = build_evidence_package_payload(job, self.store, self.uploads)
        self.assertEqual(
            payload["sbom_source_declarations"]["classification"],
            "SELF_TEST_NOT_CUSTOMER_EVIDENCE",
        )
        # Jobs without a handoff receipt carry an empty declarations dict so
        # the overview sheet and downstream consumers see nothing to render.
        payload_default = build_evidence_package_payload(self._job(), self.store, self.uploads)
        self.assertEqual(payload_default["sbom_source_declarations"], {})

    def test_actively_exploited_derivation(self) -> None:
        # actively_exploited = KEV signal OR exploited_since present (matcher rule)
        matches = [
            {"component_name": "lib", "exploitation_status": "KEV已知利用信号", "euvd_id": "E1", "match_status": "已匹配"},
            {"component_name": "xml", "exploitation_status": "未列入当前KEV快照（不代表未被利用）", "euvd_id": "E2", "match_status": "已匹配"},
            {"component_name": "net", "exploited_since": "2026-07-01", "euvd_id": "E3", "match_status": "已匹配"},
        ]
        payload = build_evidence_package_payload(self._job(matches), self.store, self.uploads)
        self.assertEqual(
            [m["actively_exploited"] for m in payload["matches"]],
            [True, False, True],  # KEV, no-signal, exploited_since
        )

    def test_no_human_decision_disclaimer(self) -> None:
        payload = build_evidence_package_payload(self._job(), self.store, self.uploads)
        self.assertEqual(payload["cra_reportable_judgments"], [])
        self.assertIn("Art.14", payload["disclaimer"])
        self.assertIn("不构成法规符合性", payload["disclaimer"])
        self.assertIn("不是外部签名", payload["disclaimer"])

    def test_unmatched_components_derived(self) -> None:
        components = [
            {"name": "matched", "version": "1", "confirmed_count": 2, "review_count": 0},
            {"name": "review", "version": "2", "confirmed_count": 0, "review_count": 1},
            {"name": "unmatched", "version": "3", "confirmed_count": 0, "review_count": 0},
        ]
        payload = build_evidence_package_payload(
            self._job(components=components), self.store, self.uploads
        )
        self.assertEqual(
            [c["name"] for c in payload["unmatched_components"]], ["unmatched"]
        )

    def test_sbom_rehash_match(self) -> None:
        content = b"sbom-bytes"
        source_sha = hashlib.sha256(content).hexdigest()
        self._write_upload("up-1", content, source_sha)
        job = self._job(source_sha=source_sha, upload_id="up-1")
        payload = build_evidence_package_payload(job, self.store, self.uploads)
        self.assertEqual(payload["sbom_integrity"]["rehash_sha256"], source_sha)
        self.assertTrue(payload["sbom_integrity"]["match"])


class EvidencePackageWriterTests(unittest.TestCase):
    """Phase B writers — _safe_cell safety + ZIP manifest integrity."""

    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        self.uploads = self.root / "uploads"
        self.uploads.mkdir()
        self.store = WorkflowStore(self.root / "wb.sqlite3")
        # adversarial affected_versions to prove _safe_cell routing
        self.payload = build_evidence_package_payload(
            {
                "id": "job-1",
                "project_name": "P",
                "result": {
                    "matches": [
                        {
                            "component_name": "lib",
                            "exploitation_status": "KEV已知利用信号",
                            "match_status": "已匹配",
                            "affected_versions": "=cmd|/c calc!A1",
                        }
                    ],
                    "components": [],
                    "summary": {},
                    "data_provenance": {},
                },
            },
            self.store,
            self.uploads,
        )

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_write_evidence_xlsx_routes_through_safe_cell(self) -> None:
        from openpyxl import load_workbook

        xlsx_bytes = write_evidence_xlsx(self.payload)
        wb = load_workbook(BytesIO(xlsx_bytes))
        self.assertIn("匹配结果", wb.sheetnames)
        # the adversarial value must be escaped (leading ' from _safe_cell), not
        # stored as a live formula
        found = False
        for row in wb["匹配结果"].iter_rows(values_only=True):
            for val in row:
                if val and "cmd" in str(val):
                    self.assertTrue(str(val).startswith("'"))
                    found = True
        self.assertTrue(found)

    def test_write_evidence_zip_manifest_integrity(self) -> None:
        import zipfile

        zip_bytes = write_evidence_package_zip(self.payload)
        self.assertTrue(zipfile.is_zipfile(BytesIO(zip_bytes)))
        with zipfile.ZipFile(BytesIO(zip_bytes)) as archive:
            names = archive.namelist()
            for required in (
                "manifest.json",
                "evidence.json",
                "evidence.xlsx",
                "audit_trail.json",
            ):
                self.assertIn(required, names)
            manifest = json.loads(archive.read("manifest.json"))
            for entry in manifest["entries"]:
                data = archive.read(entry["path"])
                self.assertEqual(
                    hashlib.sha256(data).hexdigest(), entry["sha256"]
                )
                self.assertEqual(len(data), entry["bytes"])


if __name__ == "__main__":
    unittest.main()
