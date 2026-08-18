"""Phase B assessment-support evidence bundle builder + writers.

Produces a per-job bundle (manifest + evidence.json + evidence.xlsx + audit
trail) an authorized reviewer can inspect to see what SBOM was checked, what
matched, what human decision was recorded, and whether the local hash-linked
audit chain verifies. The bundle is not an externally signed or regulator-
approved record. Read-only over job JSON + workbench.sqlite3 + uploads.
"""

from __future__ import annotations

import hashlib
import json
import uuid
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .spreadsheet_io import _safe_cell
from .workflow_store import WorkflowStore

EVIDENCE_PACKAGE_SCHEMA_VERSION = "1.0"


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def build_evidence_package_payload(
    job: dict[str, Any],
    workflow_store: WorkflowStore | None,
    uploads_dir: Path,
) -> dict[str, Any]:
    """Assemble the machine-readable evidence payload from a completed job."""
    result = job.get("result") or {}
    matches = result.get("matches") or []
    components = result.get("components") or []
    summary = result.get("summary") or {}
    data_provenance = result.get("data_provenance") or {}
    job_id = str(job.get("id") or "")

    cases = _cases_for_job(workflow_store, job_id)
    judgments = [
        _cra_judgment(c)
        for c in cases
        if (c.get("art14_decision") or "not_assessed") != "not_assessed"
    ]
    audit_trail = workflow_store.export_audit_trail(job_id) if workflow_store else []
    chain_verify = (
        workflow_store.verify_audit_chain()
        if workflow_store
        else {
            "verified": True,
            "broken_at": None,
            "event_count": 0,
            "verified_at": _utc_now(),
        }
    )

    return {
        "schema_version": EVIDENCE_PACKAGE_SCHEMA_VERSION,
        "generated_at": _utc_now(),
        "generator": "EUVD Dependency Workbench",
        "project_identity": {
            "job_id": job_id,
            "project_name": job.get("project_name") or "",
            "project_version": job.get("project_version") or "",
            "customer": job.get("customer") or "",
            "software_build": job.get("software_build") or "",
        },
        "sbom_integrity": _sbom_integrity(job, uploads_dir),
        "data_provenance": data_provenance,
        "summary": summary,
        "matches": [_match_finding(m) for m in matches],
        "unmatched_components": _unmatched_components(components),
        "cra_reportable_judgments": judgments,
        "audit_trail": audit_trail,
        "chain_verify_result": chain_verify,
        "pre7_evidence_summary": job.get("pre7_evidence_summary"),
        "sbom_source_declarations": job.get("sbom_source_declarations") or {},
        "disclaimer": (
            "KEV / actively_exploited 公开信号 ≠ CRA Art.14 reportable。reportable "
            "判定仅由人工 Art.14 案例经四眼审批（technical + compliance）后 stamp。"
            "审计链 prev_hash 全局链接，本项目首事件可能引用项目外事件，"
            "chain_verify_result 为全局校验结果，但不是外部签名或时间戳。"
            "本证据包用于评估准备，不构成法规符合性、认证、正式 SRP 提交、"
            "客户交付或发布批准。"
        ),
    }


def _cases_for_job(
    store: WorkflowStore | None, job_id: str
) -> list[dict[str, Any]]:
    if not store or not job_id:
        return []
    job_cases = [c for c in store.list_cases() if c.get("job_id") == job_id]
    full: list[dict[str, Any]] = []
    for case in job_cases:
        try:
            full.append(store.get_case(case["id"]))  # adds evidence/approvals/...
        except KeyError:
            full.append(case)
    return full


def _match_finding(m: dict[str, Any]) -> dict[str, Any]:
    exploitation = str(m.get("exploitation_status") or "")
    actively = "已知利用信号" in exploitation or bool(m.get("exploited_since"))
    cve_ids = m.get("input_cve_ids") or []
    return {
        "component_name": m.get("component_name") or m.get("name") or "",
        "component_version": m.get("component_version") or m.get("version") or "",
        "component_vendor": m.get("component_vendor") or m.get("vendor") or "",
        "cve_id": (
            cve_ids[0]
            if isinstance(cve_ids, list) and cve_ids
            else (m.get("source_identifier") or "")
        ),
        "euvd_id": m.get("euvd_id") or "",
        "euvd_url": m.get("euvd_url") or "",
        "severity": m.get("severity") or "",
        "cvss_score": m.get("cvss_score"),
        "affected_versions": m.get("affected_versions") or "",
        "match_status": m.get("match_status") or "",
        "actively_exploited": actively,
        "exploited_since": m.get("exploited_since"),
        "cra_review_required": m.get("cra_review_required"),
        "art14_readiness": m.get("art14_readiness"),
    }


def _unmatched_components(components: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "name": c.get("name") or "",
            "version": c.get("version") or "",
            "vendor": c.get("vendor") or "",
        }
        for c in components
        if not c.get("confirmed_count") and not c.get("review_count")
    ]


def _sbom_integrity(job: dict[str, Any], uploads_dir: Path) -> dict[str, Any]:
    source_sha = str(job.get("source_sha256") or "")
    upload_id = str(job.get("upload_id") or "")
    rehash: str | None = None
    record_path = uploads_dir / f"{upload_id}.upload-record.json"
    if upload_id and record_path.is_file():
        try:
            record = json.loads(record_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            record = {}
        stored = record.get("stored_file") or ""
        stored_path = (
            Path(stored) if Path(stored).is_absolute() else uploads_dir / stored
        )
        if stored and stored_path.is_file():
            try:
                rehash = _sha256_file(stored_path)
            except OSError:
                rehash = None
    return {
        "source_sha256": source_sha,
        "rehash_sha256": rehash,
        "match": (rehash == source_sha) if rehash is not None else None,
    }


def _cra_judgment(c: dict[str, Any]) -> dict[str, Any]:
    """Stamp a single Art.14 case's reportable decision + its 5-leg basis.

    Legal basis cites CRA article numbers only (no paraphrased clause text —
    the authoritative reading is the official CRA text).
    """
    evidence = c.get("evidence") or []
    approvals = c.get("approvals") or []
    reliable_rows = [
        e
        for e in evidence
        if e.get("reliable_malicious_exploitation") == "yes"
        and e.get("malicious_actor_confirmed")
        and e.get("without_permission_confirmed")
        and e.get("actual_exploitation_confirmed")
        and e.get("product_relevance")
        and e.get("sha256")
        and (e.get("source_ref") or e.get("source_url"))
    ]
    return {
        "case_id": c.get("id"),
        "finding_index": c.get("finding_index"),
        "cve_id": c.get("cve_id") or "",
        "euvd_id": c.get("euvd_id") or "",
        "component_name": c.get("component_name") or "",
        "art14_decision": c.get("art14_decision") or "not_assessed",
        "workflow_status": c.get("workflow_status") or "",
        "decision_rationale": c.get("decision_rationale") or "",
        "basis": {
            "applicability": {
                "status": c.get("applicability_status"),
                "justification": c.get("applicability_justification"),
            },
            "exploitation_evidence": {
                "status": c.get("exploitation_evidence_status"),
                "summary": c.get("exploitation_evidence_summary"),
            },
            "awareness": {
                "awareness_at": c.get("awareness_at"),
                "confirmed_by": c.get("awareness_confirmed_by"),
                "basis": c.get("awareness_basis"),
            },
            "reliable_evidence_rows": len(reliable_rows),
            "four_eye_approval": {
                "technical_reviewer": c.get("technical_reviewer"),
                "compliance_reviewer": c.get("compliance_reviewer"),
                "approval_rows": len(approvals),
            },
        },
        "legal_basis": (
            "CRA Art.13/14 — Art.3(8) known vulnerability; Art.3(42) actively "
            "exploited; Art.14(1) reliable evidence of active exploitation + "
            "24h awareness clock. (条款号引用；条文释义以官方文本为准。)"
        ),
    }


# --- Writers (Phase B B3) ---


def write_evidence_xlsx(payload: dict[str, Any]) -> bytes:
    """Human-readable workbook (4 sheets). Every cell passes through _safe_cell
    (csv-injection =/+/-/@ escaping + XML control-char stripping)."""
    workbook = Workbook()
    _overview_sheet(workbook.active, payload)
    _table_sheet(
        workbook.create_sheet("匹配结果"),
        ["组件", "版本", "CVE", "EUVD", "严重度", "CVSS", "受影响版本", "匹配状态", "积极利用", "cra_review_required"],
        _match_rows(payload.get("matches") or []),
    )
    _table_sheet(
        workbook.create_sheet("CRA报告判定"),
        ["案件", "CVE", "EUVD", "art14_decision", "workflow_status", "可靠证据行", "四眼审批行"],
        _judgment_rows(payload.get("cra_reportable_judgments") or []),
    )
    _table_sheet(
        workbook.create_sheet("审计链"),
        ["id", "case_id", "event_type", "actor", "created_at", "prev_hash", "payload_sha256"],
        _audit_rows(payload.get("audit_trail") or []),
    )
    buf = BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _overview_sheet(ws: Any, payload: dict[str, Any]) -> None:
    pi = payload.get("project_identity") or {}
    si = payload.get("sbom_integrity") or {}
    dp = payload.get("data_provenance") or {}
    sm = payload.get("summary") or {}
    cv = payload.get("chain_verify_result") or {}
    rows = [
        ("EUVD 评估支持证据包（非符合性结论）", None),
        ("项目", pi.get("project_name")),
        ("版本", pi.get("project_version")),
        ("客户", pi.get("customer")),
        ("软件构建", pi.get("software_build")),
        ("job_id", pi.get("job_id")),
        ("生成时间(UTC)", payload.get("generated_at")),
        ("SBOM source_sha256", si.get("source_sha256")),
        ("SBOM rehash_sha256", si.get("rehash_sha256")),
        ("SBOM 完整性匹配", si.get("match")),
        ("EUVD 数据截止", dp.get("last_successful_to_date")),
        ("快照 sha256", dp.get("snapshot_sha256")),
        ("参考数据新鲜度", dp.get("reference_data_freshness")),
        ("组件数", sm.get("component_count")),
        ("已确认发现", sm.get("confirmed_findings")),
        ("需复核发现", sm.get("review_findings")),
        ("KEV 发现", sm.get("known_exploited_findings")),
        ("审计链校验", "通过" if cv.get("verified") else f"断裂@{cv.get('broken_at')}"),
        ("审计事件数", cv.get("event_count")),
        ("免责声明", payload.get("disclaimer")),
    ]
    declarations = payload.get("sbom_source_declarations") or {}
    if declarations:
        # Surface the candidate-SBOM boundary right after the integrity check
        # so an authorized reviewer sees classification/provenance
        # alongside the hash, before any CRA-reportable judgment.
        insert_at = next(
            (i for i, (k, _) in enumerate(rows) if k == "SBOM 完整性匹配"),
            len(rows) - 1,
        ) + 1
        rows[insert_at:insert_at] = [
            ("SBOM 来源声明", None),
            ("SBOM 分类", declarations.get("classification")),
            ("SBOM provenance", declarations.get("source_binding_status")),
            ("SBOM 单向契约", declarations.get("direction")),
            ("SBOM 权威边界", declarations.get("authority_boundary")),
        ]
    for i, (key, val) in enumerate(rows, start=1):
        c1 = ws.cell(row=i, column=1, value=_safe_cell(key))
        c1.font = Font(bold=True)
        ws.cell(row=i, column=2, value=_safe_cell(val))


def _table_sheet(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=_safe_cell(header))
        cell.font = Font(bold=True)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=_safe_cell(val))


def _match_rows(matches: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            m.get("component_name"),
            m.get("component_version"),
            m.get("cve_id"),
            m.get("euvd_id"),
            m.get("severity"),
            m.get("cvss_score"),
            m.get("affected_versions"),
            m.get("match_status"),
            "是" if m.get("actively_exploited") else "",
            m.get("cra_review_required"),
        ]
        for m in matches
    ]


def _judgment_rows(judgments: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            j.get("case_id"),
            j.get("cve_id"),
            j.get("euvd_id"),
            j.get("art14_decision"),
            j.get("workflow_status"),
            (j.get("basis") or {}).get("reliable_evidence_rows"),
            ((j.get("basis") or {}).get("four_eye_approval") or {}).get("approval_rows"),
        ]
        for j in judgments
    ]


def _audit_rows(audit: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            e.get("id"),
            e.get("case_id"),
            e.get("event_type"),
            e.get("actor"),
            e.get("created_at"),
            (e.get("prev_hash") or "")[:16],
            (e.get("payload_sha256") or "")[:16],
        ]
        for e in audit
    ]


def write_evidence_package_zip(payload: dict[str, Any]) -> bytes:
    """Assemble the ZIP bundle: manifest.json + evidence.json + evidence.xlsx +
    audit_trail.json. Per-entry sha256 stamped in manifest for tamper check."""
    entries: dict[str, bytes] = {}
    entries["evidence.json"] = json.dumps(
        payload, ensure_ascii=False, indent=2
    ).encode("utf-8")
    entries["evidence.xlsx"] = write_evidence_xlsx(payload)
    entries["audit_trail.json"] = json.dumps(
        {
            "chain_verify_result": payload.get("chain_verify_result"),
            "audit_trail": payload.get("audit_trail"),
            "scoping_note": (
                "prev_hash 全局链接；本项目首事件可能引用项目外事件。"
                "chain_verify_result 为全局校验结果。"
            ),
        },
        ensure_ascii=False,
        indent=2,
    ).encode("utf-8")
    manifest = _build_manifest(payload, entries)
    entries["manifest.json"] = json.dumps(
        manifest, ensure_ascii=False, indent=2
    ).encode("utf-8")
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in entries.items():
            archive.writestr(name, data)
    return buf.getvalue()


def _build_manifest(
    payload: dict[str, Any], entries: dict[str, bytes]
) -> dict[str, Any]:
    return {
        "bundle_id": str(uuid.uuid4()),
        "schema_version": EVIDENCE_PACKAGE_SCHEMA_VERSION,
        "job_id": (payload.get("project_identity") or {}).get("job_id"),
        "generated_at": payload.get("generated_at"),
        "generator": payload.get("generator"),
        "entries": [
            {
                "path": name,
                "sha256": hashlib.sha256(data).hexdigest(),
                "bytes": len(data),
            }
            for name, data in entries.items()
        ],
    }
