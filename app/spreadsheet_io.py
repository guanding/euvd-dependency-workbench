from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from charset_normalizer import from_bytes
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .matcher import (
    Component,
    enrich_component,
    extract_identifiers,
    normalize_key,
    repair_text,
)
from .sbom_evidence import extract_cyclonedx_evidence
from .template_builder import write_public_template


HEADER_ALIASES = {
    "name": {
        "component",
        "componentname",
        "name",
        "product",
        "productname",
        "software",
        "packagename",
        "组件",
        "组件名称",
        "软件名称",
        "产品名称",
        "包名",
        "component name 组件名称",
    },
    "version": {
        "version",
        "componentversion",
        "productversion",
        "版本",
        "组件版本",
        "软件版本",
        "产品版本",
        "component version 组件版本",
    },
    "vendor": {
        "vendor",
        "supplier",
        "manufacturer",
        "publisher",
        "author",
        "group",
        "namespace",
        "厂商",
        "供应商",
        "制造商",
        "发布者",
        "组织",
        "component producer 组件生产者",
    },
    "purl": {
        "purl",
        "packageurl",
        "package-url",
        "包地址",
        "purl package url",
    },
    "cpe": {
        "cpe",
        "cpe22",
        "cpe23",
        "cpe2.3",
        "通用平台枚举",
    },
    "scope": {
        "scope",
        "type",
        "componenttype",
        "范围",
        "类型",
        "组件类型",
        "component category 组件类别",
    },
    "license": {
        "license",
        "licenses",
        "许可证",
        "许可",
        "license 许可证",
    },
    "cve": {
        "cve",
        "cveid",
        "cveids",
        "vulnerability",
        "vulnerabilityid",
        "漏洞",
        "漏洞编号",
        "cve编号",
        "cve 漏洞编号",
    },
    "euvd": {
        "euvd",
        "euvdid",
        "euvdids",
        "euvd编号",
        "enisaid",
        "euvd id euvd编号",
    },
}


# D1: customer-template Metadata-sheet labels. The 01_Metadata_元数据 sheet is a
# label/value table (A=label, B=customer entry). These map labels to product
# identity (prefilled into the operator's inputs) plus extra fields retained only
# as audit evidence. Matched by substring on normalize_key() so
# "Product name 产品名称", "产品名称", and "Product name" all resolve.
METADATA_LABEL_ALIASES = {
    "product_name": {
        "product name 产品名称",
        "产品名称",
        "product name",
        "productname",
    },
    "product_version": {
        "product version 产品版本",
        "产品版本",
        "product version",
        "productversion",
    },
    "software_build": {
        "build id 构建号",
        "构建号",
        "build id",
        "build identifier",
    },
    "hardware_revision": {
        "hardware revision 硬件修订",
        "硬件修订",
        "hardware revision",
    },
    "release_date": {
        "release date 发布日期",
        "发布日期",
        "release date",
    },
    "sbom_version": {
        "sbom version sbom版本",
        "sbom版本",
        "sbom version",
    },
}
METADATA_IDENTITY_FIELDS = ("product_name", "product_version", "software_build")


def _match_metadata_label(label: Any) -> str | None:
    text = normalize_key(label)
    if not text:
        return None
    for field, aliases in METADATA_LABEL_ALIASES.items():
        for alias in aliases:
            # Exact match on the normalized label/alias. A substring test
            # over-matched (e.g. "Hardware Build ID" -> build id ->
            # software_build; "productnamespace" -> productname). A customer who
            # varies the label wording should add it to METADATA_LABEL_ALIASES
            # deliberately rather than rely on loose substring matching.
            if normalize_key(alias) == text:
                return field
    return None


def _read_label_value_pairs(worksheet: Any, max_rows: int = 30) -> list[tuple[Any, Any]]:
    pairs: list[tuple[Any, Any]] = []
    for row in worksheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
        cells = list(row) + [None, None]
        label = cells[0]
        value = cells[1]
        if label is None and value is None:
            continue
        pairs.append((label, value))
    return pairs


def _looks_like_metadata(worksheet: Any) -> bool:
    matched = {
        _match_metadata_label(label)
        for label, _ in _read_label_value_pairs(worksheet, max_rows=15)
    }
    return len(matched & set(METADATA_IDENTITY_FIELDS)) >= 2


def _extract_metadata_binding(workbook: Any) -> dict[str, Any] | None:
    """Extract product identity from a customer Metadata sheet.

    Identifies the sheet by name (contains "metadata"/"元数据") with a content
    fallback (>=2 product-identity labels), reads its label/value column, and
    maps labels to prefilled identity fields plus audit-only evidence. Returns
    None when no Metadata sheet is present (csv/json, or xlsx without one).
    """
    visible = [s for s in workbook.worksheets if s.sheet_state == "visible"] or [
        workbook.active
    ]
    metadata_sheet = None
    for sheet in visible:
        title = sheet.title or ""
        # Token match (split on separators) so a sheet literally named
        # "mymetadatanotes" does not match, but "01_Metadata_元数据" does.
        title_tokens = re.split(r"[\s_\-/]+", title.casefold())
        if "metadata" in title_tokens or "元数据" in title:
            metadata_sheet = sheet
            break
    if metadata_sheet is None:
        for sheet in visible:
            if _looks_like_metadata(sheet):
                metadata_sheet = sheet
                break
    if metadata_sheet is None:
        return None
    fields: dict[str, str] = {}
    evidence: dict[str, str] = {}
    raw: dict[str, str] = {}
    for label, value in _read_label_value_pairs(metadata_sheet, max_rows=30):
        text = repair_text(value)
        label_text = repair_text(label)
        if label_text:
            raw[label_text] = text
        field = _match_metadata_label(label)
        if not field or not text:
            continue
        if field in METADATA_IDENTITY_FIELDS:
            fields.setdefault(field, text)
        else:
            evidence.setdefault(field, text)
    return {
        "fields": fields,
        "evidence": evidence,
        "raw": raw,
        "source_sheet": metadata_sheet.title,
    }


# SBOM Workbench handoff receipt boundary constants. These must match the
# corresponding constants in euvd_handoff.py (offline-sbom-evidence-workbench);
# any mismatch means the receipt is not a valid one-way handoff and must be
# rejected fail-closed rather than silently accepted as a product SBOM.
HANDOFF_CLASSIFICATION = "SELF_TEST_NOT_CUSTOMER_EVIDENCE"
HANDOFF_AUTHORITY_BOUNDARY = "NO_SBOM_FACT_RELEASE_CONFORMITY_OR_REPORTING_AUTHORITY"
HANDOFF_KEV_BOUNDARY = "KEV_PRESENCE_IS_PRIORITIZATION_ONLY_ABSENCE_IS_NOT_NON_EXPLOITATION_PROOF"
HANDOFF_DIRECTION = "SBOM_TO_EUVD_ONLY"
# Receipt fields carried forward as audit-only evidence. None is a product
# identity field, so `fields` stays empty and the prefill chain is unaffected.
# source_run_id is the workbench's internal run id, not a customer identifier.
HANDOFF_EVIDENCE_KEYS = (
    "classification",
    "authority_boundary",
    "kev_boundary",
    "direction",
    "reverse_fact_write",
    "source_binding_status",
    "source_profile_id",
    "cyclonedx_sha256",
    "component_record_count",
    "purl_coverage",
    "version_coverage",
    "handoff_id",
    "schema_version",
    "source_run_id",
)


def _extract_handoff_binding(cyclonedx_path: Path) -> dict[str, Any] | None:
    """Read a co-located SBOM Workbench receipt (``{stem}.receipt.json``) and
    return a ``metadata_binding`` shaped like ``_extract_metadata_binding``:
    boundary declarations go to ``evidence`` (audit-only), ``fields`` is empty
    (a receipt carries no product identity), and ``source_sheet`` is
    ``"receipt.json"`` so the existing prefill hint reads naturally.

    Returns ``None`` when no receipt sits next to the CycloneDX file (plain
    upload, behaviour unchanged). Raises ``ValueError`` on any boundary
    mismatch so the caller (``upload_preview``) turns it into HTTP 400.
    """
    receipt_path = cyclonedx_path.with_name(f"{cyclonedx_path.stem}.receipt.json")
    if not receipt_path.is_file():
        return None
    try:
        receipt = json.loads(receipt_path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"receipt.json 解析失败: {exc}") from exc
    if not isinstance(receipt, dict):
        raise ValueError("receipt.json 必须是 JSON 对象")
    if receipt.get("direction") != HANDOFF_DIRECTION:
        raise ValueError("receipt direction 必须为 SBOM_TO_EUVD_ONLY")
    if receipt.get("reverse_fact_write") is not False:
        raise ValueError("receipt reverse_fact_write 必须为 false")
    if receipt.get("classification") != HANDOFF_CLASSIFICATION:
        raise ValueError("receipt classification 不符合 SELF_TEST 边界")
    if receipt.get("authority_boundary") != HANDOFF_AUTHORITY_BOUNDARY:
        raise ValueError("receipt authority_boundary 与既定边界不符")
    if receipt.get("kev_boundary") != HANDOFF_KEV_BOUNDARY:
        raise ValueError("receipt kev_boundary 与既定边界不符")
    try:
        actual_sha256 = hashlib.sha256(cyclonedx_path.read_bytes()).hexdigest()
    except OSError as exc:  # pragma: no cover - filesystem race
        raise ValueError(f"无法读取 cyclonedx 计算 sha256: {exc}") from exc
    if receipt.get("cyclonedx_sha256") != actual_sha256:
        raise ValueError("receipt cyclonedx_sha256 与上传的 cyclonedx 不匹配")
    evidence = {key: receipt[key] for key in HANDOFF_EVIDENCE_KEYS if key in receipt}
    return {
        "fields": {},
        "evidence": evidence,
        "raw": dict(receipt),
        "source_sheet": "receipt.json",
    }


# Resource bounds: cap untrusted parsing to prevent OOM on crafted inputs.
_MAX_PARSE_ROWS_PER_SHEET = 200_000
_MAX_VULNERABILITIES = 50_000

_ILLEGAL_XML_CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool, datetime)):
        return value
    text = repair_text(value)
    # Strip XML-illegal control characters (NUL, 0x01-0x08, 0x0B, 0x0C, 0x0E-0x1F).
    # openpyxl raises IllegalCharacterError on them, which would abort the whole
    # report write after the long match run has already completed.
    text = _ILLEGAL_XML_CONTROL_RE.sub("", text)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _dedupe_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = repair_text(value) or f"Column {index}"
        if base not in seen:
            seen[base] = 1
            headers.append(base)
            continue
        seen[base] += 1
        # Use a '__N' suffix (not '(N)') so it cannot collide with a source
        # header already shaped like "X (1)".
        deduped = f"{base}__{seen[base]}"
        while deduped in seen:
            seen[base] += 1
            deduped = f"{base}__{seen[base]}"
        seen[deduped] = 1
        headers.append(deduped)
    return headers


def infer_mapping(headers: list[str], rows: list[dict[str, Any]] | None = None) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field, aliases in HEADER_ALIASES.items():
        alias_keys = {normalize_key(alias) for alias in aliases}
        candidates = [h for h in headers if normalize_key(h) in alias_keys]
        if not candidates:
            continue
        if rows:
            # Prefer a candidate column that actually has data; an empty
            # first-match column (e.g. a blank template column) would otherwise
            # be mapped and yield zero components.
            populated = [
                h
                for h in candidates
                if any(repair_text(r.get(h)) for r in rows[:200])
            ]
            mapping[field] = populated[0] if populated else candidates[0]
        else:
            mapping[field] = candidates[0]
    return mapping


def _find_header_row(rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1
    all_aliases = {
        normalize_key(alias)
        for aliases in HEADER_ALIASES.values()
        for alias in aliases
    }
    for index, row in enumerate(rows[:50]):
        keys = {normalize_key(value) for value in row if repair_text(value)}
        score = len(keys & all_aliases)
        if any(key in {normalize_key(alias) for alias in HEADER_ALIASES["name"]} for key in keys):
            # A name-alias hit is a header signal, but the old +3 could let a
            # data row that repeats an alias word outrank the real header. +1
            # keeps it a tie-breaker, not an override.
            score += 1
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _read_xlsx(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    candidates: list[
        tuple[tuple[int, int, int, int], Any, list[list[Any]], int, list[str]]
    ] = []
    try:
        visible_sheets = [
            sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"
        ] or [workbook.active]
        for order, worksheet in enumerate(visible_sheets):
            sheet_rows: list[list[Any]] = []
            for row in worksheet.iter_rows(values_only=True):
                sheet_rows.append(list(row))
                if len(sheet_rows) >= _MAX_PARSE_ROWS_PER_SHEET:
                    break
            if not sheet_rows:
                continue
            sheet_header_index = _find_header_row(sheet_rows)
            sheet_headers = _dedupe_headers(sheet_rows[sheet_header_index])
            sheet_mapping = infer_mapping(sheet_headers)
            identity_headers = [
                sheet_mapping.get(field)
                for field in ("name", "purl", "cpe", "cve", "euvd")
                if sheet_mapping.get(field)
            ]
            populated_identity_rows = sum(
                1
                for row in sheet_rows[sheet_header_index + 1 :]
                if any(
                    repair_text(row[sheet_headers.index(header)])
                    for header in identity_headers
                    if sheet_headers.index(header) < len(row)
                )
            )
            # Prefer sheets that actually contain identity data (populated rows)
            # over sheets with more identity columns but no rows (e.g. an empty
            # template sheet that lists every alias header).
            score = (
                populated_identity_rows,
                len(identity_headers),
                len(sheet_mapping),
                -order,
            )
            candidates.append(
                (
                    score,
                    worksheet,
                    sheet_rows,
                    sheet_header_index,
                    sheet_headers,
                )
            )
        if not candidates:
            raise ValueError("工作簿没有可读取的数据")
        _, worksheet, raw_rows, header_index, headers = max(
            candidates, key=lambda item: item[0]
        )
        worksheet_title = worksheet.title
        metadata_binding = _extract_metadata_binding(workbook)
    finally:
        workbook.close()
    # Ragged rows: if any data row is wider than the header, extend headers with
    # extra_N columns so the trailing cells are kept instead of silently dropped.
    max_width = max(
        (len(raw) for raw in raw_rows[header_index + 1 :]), default=len(headers)
    )
    if max_width > len(headers):
        extra_count = max_width - len(headers)
        headers = headers + [f"extra_{i + 1}" for i in range(extra_count)]
    rows: list[dict[str, Any]] = []
    for raw in raw_rows[header_index + 1 :]:
        padded = list(raw) + [None] * max(0, len(headers) - len(raw))
        record = {headers[index]: padded[index] for index in range(len(headers))}
        if any(repair_text(value) for value in record.values()):
            rows.append(record)
    return {
        "kind": "table",
        "sheet": worksheet_title,
        "header_row": header_index + 1,
        "headers": headers,
        "rows": rows,
        "metadata_binding": metadata_binding,
    }


def _decode_csv(path: Path) -> str:
    data = path.read_bytes()
    # Encoding detection is a declared, pinned runtime dependency. Making it
    # optional caused clean containers to take the permissive gb18030 fallback
    # and silently turn cp1252 names (for example Müller) into CJK gibberish.
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        return data.decode("utf-16")
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass

    candidates = from_bytes(data)
    # Short GB18030 CSV headers can tie with Big5, CP949 and EUC-JIS at zero
    # chaos. This product has historically promised GB18030 input, not those
    # other legacy encodings, so select the supported candidate explicitly.
    for candidate in candidates:
        encoding = (candidate.encoding or "").lower().replace("_", "-")
        if encoding in {"gb18030", "gbk", "gb2312"}:
            return str(candidate)
    best = candidates.best()
    if best is not None and best.encoding:
        return str(best)
    for encoding in ("gb18030", "cp1252"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _read_csv(path: Path) -> dict[str, Any]:
    text = _decode_csv(path)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    raw_rows = [row for row in reader]
    if not raw_rows:
        raise ValueError("CSV 没有可读取的数据")
    header_index = _find_header_row(raw_rows)
    headers = _dedupe_headers(raw_rows[header_index])
    max_width = max(
        (len(raw) for raw in raw_rows[header_index + 1 :]), default=len(headers)
    )
    if max_width > len(headers):
        extra_count = max_width - len(headers)
        headers = headers + [f"extra_{i + 1}" for i in range(extra_count)]
    rows = []
    for raw in raw_rows[header_index + 1 :]:
        padded = list(raw) + [""] * max(0, len(headers) - len(raw))
        record = {headers[index]: padded[index] for index in range(len(headers))}
        if any(repair_text(value) for value in record.values()):
            rows.append(record)
    return {
        "kind": "table",
        "sheet": "CSV",
        "header_row": header_index + 1,
        "headers": headers,
        "rows": rows,
    }


def _license_text(component: dict[str, Any]) -> str:
    values: list[str] = []
    for entry in component.get("licenses") or []:
        license_data = entry.get("license") or {}
        value = license_data.get("id") or license_data.get("name") or entry.get("expression")
        if value:
            values.append(repair_text(value))
    return ", ".join(values)


def _flatten_cyclonedx_components(components: list[Any]) -> list[dict[str, Any]]:
    flattened: list[dict[str, Any]] = []
    stack: list[tuple[Any, int]] = [
        (component, 1) for component in reversed(components)
    ]
    while stack:
        component, depth = stack.pop()
        if not isinstance(component, dict):
            continue
        if len(flattened) >= 10_000:
            raise ValueError("CycloneDX 组件数超过解析安全上限 10000")
        if depth > 100:
            raise ValueError("CycloneDX 组件嵌套深度超过解析安全上限 100")
        flattened.append(component)
        children = component.get("components")
        if isinstance(children, list):
            stack.extend((child, depth + 1) for child in reversed(children))
    return flattened


def _read_cyclonedx(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if payload.get("bomFormat") != "CycloneDX" or not isinstance(payload.get("components"), list):
        raise ValueError("JSON 不是可识别的 CycloneDX SBOM")
    preserved = extract_cyclonedx_evidence(payload)
    warnings = list(preserved["warnings"])
    flattened_components = _flatten_cyclonedx_components(payload["components"])
    headers = ["组件名称", "版本", "厂商", "PURL", "CPE", "类型", "许可证", "CVE", "EUVD ID"]
    vulnerability_ids: dict[str, dict[str, list[str]]] = {}
    vulnerabilities = payload.get("vulnerabilities") or []
    if not isinstance(vulnerabilities, list):
        warnings.append("CycloneDX vulnerabilities 不是数组，已跳过漏洞关联解析")
        vulnerabilities = []
    if len(vulnerabilities) > _MAX_VULNERABILITIES:
        warnings.append(
            f"CycloneDX vulnerabilities 超过解析上限 {_MAX_VULNERABILITIES}，"
            f"已截断前 {_MAX_VULNERABILITIES} 条"
        )
        vulnerabilities = vulnerabilities[:_MAX_VULNERABILITIES]
    component_refs = {
        repair_text(component.get("bom-ref"))
        for component in flattened_components
        if isinstance(component, dict) and repair_text(component.get("bom-ref"))
    }
    for vulnerability_index, vulnerability in enumerate(vulnerabilities):
        if not isinstance(vulnerability, dict):
            warnings.append(
                f"CycloneDX vulnerabilities[{vulnerability_index}] 不是对象，已跳过"
            )
            continue
        identifier = repair_text(vulnerability.get("id")).upper()
        if not identifier:
            continue
        kind = "euvd" if identifier.startswith("EUVD-") else "cve" if identifier.startswith("CVE-") else ""
        if not kind:
            continue
        affects = vulnerability.get("affects") or []
        if not isinstance(affects, list):
            warnings.append(
                f"CycloneDX vulnerabilities[{vulnerability_index}].affects 不是数组，已跳过"
            )
            continue
        for affected_index, affected in enumerate(affects):
            if not isinstance(affected, dict):
                warnings.append(
                    "CycloneDX "
                    f"vulnerabilities[{vulnerability_index}].affects[{affected_index}] "
                    "不是对象，已跳过"
                )
                continue
            reference = repair_text(affected.get("ref"))
            if not reference:
                continue
            if reference not in component_refs:
                warnings.append(
                    f"{identifier} 的 affects ref 未在顶层 components 中找到: {reference}"
                )
            vulnerability_ids.setdefault(reference, {"cve": [], "euvd": []})[kind].append(identifier)
    rows = []
    for component in flattened_components:
        if not isinstance(component, dict):
            warnings.append("CycloneDX components 包含非对象条目，已跳过")
            continue
        supplier = component.get("supplier") or {}
        vendor = (
            supplier.get("name") if isinstance(supplier, dict) else ""
        ) or component.get("publisher") or component.get("group") or ""
        linked = vulnerability_ids.get(repair_text(component.get("bom-ref")), {"cve": [], "euvd": []})
        rows.append(
            {
                "组件名称": component.get("name") or "",
                "版本": component.get("version") or "",
                "厂商": vendor,
                "PURL": component.get("purl") or "",
                "CPE": component.get("cpe") or "",
                "类型": component.get("type") or "",
                "许可证": _license_text(component),
                "CVE": ", ".join(dict.fromkeys(linked["cve"])),
                "EUVD ID": ", ".join(dict.fromkeys(linked["euvd"])),
            }
        )
    result = {
        "kind": "cyclonedx",
        "sheet": "CycloneDX",
        "header_row": 1,
        "spec_version": repair_text(payload.get("specVersion")),
        "serial_number": repair_text(payload.get("serialNumber")),
        "bom_version": payload.get("version"),
        "headers": headers,
        "rows": rows,
        "mapping": {
            "name": "组件名称",
            "version": "版本",
            "vendor": "厂商",
            "purl": "PURL",
            "cpe": "CPE",
            "scope": "类型",
            "license": "许可证",
            "cve": "CVE",
            "euvd": "EUVD ID",
        },
    }
    result.update(preserved)
    result["warnings"] = list(dict.fromkeys(warnings))
    return result


def read_sbom(path: Path) -> dict[str, Any]:
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        result = _read_xlsx(path)
    elif suffix in {".csv", ".tsv"}:
        result = _read_csv(path)
    elif suffix == ".json":
        result = _read_cyclonedx(path)
        handoff_binding = _extract_handoff_binding(path)
        if handoff_binding is not None:
            result["metadata_binding"] = handoff_binding
    else:
        raise ValueError("仅支持 .xlsx、.xlsm、.csv、.tsv 和 CycloneDX .json")
    result.setdefault("mapping", infer_mapping(result["headers"], result["rows"]))
    identity_headers = [
        result["mapping"].get(field)
        for field in ("name", "purl", "cpe", "cve", "euvd")
        if result["mapping"].get(field)
    ]
    if identity_headers:
        result["rows"] = [
            row
            for row in result["rows"]
            if any(repair_text(row.get(header)) for header in identity_headers)
        ]
    result.setdefault("metadata_binding", None)
    return result


def build_components(parsed: dict[str, Any], mapping: dict[str, str], max_components: int) -> list[Component]:
    name_header = mapping.get("name")
    if not (name_header or mapping.get("cve") or mapping.get("euvd")):
        raise ValueError("必须选择“组件名称”、CVE 或 EUVD ID 列之一")

    components: list[Component] = []
    seen: set[tuple[str, str, str, str, str, str]] = set()
    for offset, row in enumerate(parsed["rows"], start=parsed["header_row"] + 1):
        component = enrich_component(
            Component(
                row_number=offset,
                name=repair_text(row.get(name_header)),
                version=repair_text(row.get(mapping.get("version", ""))),
                vendor=repair_text(row.get(mapping.get("vendor", ""))),
                purl=repair_text(row.get(mapping.get("purl", ""))),
                cpe=repair_text(row.get(mapping.get("cpe", ""))),
                scope=repair_text(row.get(mapping.get("scope", ""))),
                license=repair_text(row.get(mapping.get("license", ""))),
                cve_ids=", ".join(
                    extract_identifiers(row.get(mapping.get("cve", "")), "cve")
                ),
                euvd_ids=", ".join(
                    extract_identifiers(row.get(mapping.get("euvd", "")), "euvd")
                ),
            )
        )
        if not component.name and not (component.cve_ids or component.euvd_ids):
            continue
        if not component.name:
            component.name = component.cve_ids or component.euvd_ids
        identity = (
            normalize_key(component.name),
            normalize_key(component.version),
            normalize_key(component.vendor),
            normalize_key(component.purl or component.cpe),
            normalize_key(component.cve_ids),
            normalize_key(component.euvd_ids),
        )
        if identity in seen:
            continue
        seen.add(identity)
        components.append(component)
        if len(components) > max_components:
            raise ValueError(f"组件数量超过限制 {max_components}")
    if not components:
        raise ValueError("没有识别到有效组件，请检查列映射")
    return components


THIN = Side(style="thin", color="D8DEE8")
HEADER_FILL = PatternFill("solid", fgColor="233247")
SUBTLE_FILL = PatternFill("solid", fgColor="EAF0F6")
CYAN_FILL = PatternFill("solid", fgColor="D8F3F7")
GREEN_FILL = PatternFill("solid", fgColor="DDF3E4")
YELLOW_FILL = PatternFill("solid", fgColor="FFF0C2")
RED_FILL = PatternFill("solid", fgColor="FBE0E0")


def _style_header(row_cells: list[Any]) -> None:
    for cell in row_cells:
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def _add_table(worksheet: Any, name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _fit_columns(worksheet: Any, max_width: int = 52) -> None:
    for column_cells in worksheet.columns:
        width = 8
        for cell in list(column_cells)[:200]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _write_rows(worksheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    # Sanitize headers too (not just body cells): customer-controlled CSV/XLSX
    # headers reach the 输入SBOM快照 sheet and a '=' prefix would become a live
    # formula in the report (csv-injection).
    worksheet.append([_safe_cell(header) for header in headers])
    _style_header(list(worksheet[1]))
    for row in rows:
        worksheet.append([_safe_cell(row.get(header, "")) for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _fit_columns(worksheet)


def write_report(
    output_path: Path,
    source_name: str,
    results: dict[str, Any],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总"
    summary_sheet.sheet_view.showGridLines = False

    summary_sheet["A1"] = "EUVD SBOM 漏洞匹配报告"
    summary_sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    summary_sheet["A1"].fill = HEADER_FILL
    summary_sheet.merge_cells("A1:F2")
    summary_sheet["A1"].alignment = Alignment(vertical="center")

    summary_sheet["A4"] = "输入文件"
    summary_sheet["B4"] = _safe_cell(source_name)
    summary_sheet["A5"] = "生成时间"
    summary_sheet["B5"] = (
        datetime.now(timezone.utc)
        .astimezone()
        .replace(microsecond=0, tzinfo=None)
    )
    summary_sheet["B5"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary_sheet["A6"] = "漏洞数据源"
    summary_sheet["B6"] = "ENISA European Vulnerability Database (EUVD)"
    summary_sheet["A7"] = "API"
    summary_sheet["B7"] = "https://euvdservices.enisa.europa.eu"
    summary_sheet["B7"].hyperlink = "https://euvdservices.enisa.europa.eu"
    provenance = results.get("data_provenance") or {}
    summary_sheet["A8"] = "查询模式"
    summary_sheet["B8"] = _safe_cell(
        provenance.get("query_mode") or "network-euvd-api"
    )
    summary_sheet["A9"] = "EUVD 数据截止"
    summary_sheet["B9"] = _safe_cell(
        provenance.get("last_successful_to_date") or "网络请求；无本地高水位"
    )
    summary_sheet["A10"] = "参考数据新鲜度"
    summary_sheet["B10"] = _safe_cell(
        provenance.get("reference_data_freshness") or "unknown"
    )
    summary_sheet["A11"] = "本地镜像源 SHA-256"
    summary_sheet["B11"] = _safe_cell(
        provenance.get("source_db_sha256") or "不适用"
    )

    metrics = [
        ("组件总数", results["summary"]["component_count"], CYAN_FILL),
        ("EUVD记录匹配", results["summary"]["confirmed_findings"], GREEN_FILL),
        ("KEV已知利用信号", results["summary"].get("known_exploited_findings", 0), RED_FILL),
        ("Art.14紧急人工评估", results["summary"].get("art14_review_findings", 0), RED_FILL),
        ("需人工复核", results["summary"]["review_findings"], YELLOW_FILL),
        ("未映射标识符", results["summary"].get("unmapped_identifier_count", 0), YELLOW_FILL),
        ("未发现EUVD记录的组件", results["summary"]["unmatched_components"], SUBTLE_FILL),
        ("查询错误", results["summary"]["error_count"], RED_FILL),
        ("身份信息覆盖率", f"{results['summary'].get('identity_coverage_percent', 0)}%", CYAN_FILL),
        ("EUVD 查询成功率", f"{results['summary'].get('query_coverage_percent', 0)}%", GREEN_FILL),
        ("完整分页覆盖率", f"{results['summary'].get('retrieval_coverage_percent', 0)}%", CYAN_FILL),
        ("分页截断组件", results["summary"].get("truncated_queries", 0), YELLOW_FILL),
    ]
    for index, (label, value, fill) in enumerate(metrics, start=4):
        summary_sheet.cell(index, 4, label)
        summary_sheet.cell(index, 5, value)
        summary_sheet.cell(index, 4).fill = fill
        summary_sheet.cell(index, 5).fill = fill
        summary_sheet.cell(index, 4).font = Font(bold=True)
        summary_sheet.cell(index, 5).font = Font(size=15, bold=True)
        summary_sheet.cell(index, 5).alignment = Alignment(horizontal="right")

    summary_sheet["A18"] = "结论使用说明"
    summary_sheet["A18"].font = Font(bold=True)
    summary_sheet["A18"].fill = SUBTLE_FILL
    summary_sheet.merge_cells("A18:F18")
    summary_sheet["A19"] = (
        "“EUVD精确匹配”仅确认输入 CVE/EUVD 与公开 EUVD 记录的映射。"
        "“KEV已知利用信号”表示 EU/CISA KEV 已列入，不自动证明客户产品已满足 CRA Art.14。"
        "最终判定仍需核验客户产品是否包含受影响代码、版本/配置/可达性、VEX、产品级恶意利用可靠证据，"
        "并由制造商人工确认 awareness 时间。未列入 KEV 不代表未被利用；SRP 是报告入口，不是查询数据库。"
    )
    summary_sheet.merge_cells("A19:F23")
    summary_sheet["A19"].alignment = Alignment(vertical="top", wrap_text=True)
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 42
    summary_sheet.column_dimensions["C"].width = 3
    summary_sheet.column_dimensions["D"].width = 24
    summary_sheet.column_dimensions["E"].width = 16
    summary_sheet.column_dimensions["F"].width = 3

    result_headers = [
        "mapping_status",
        "source_identifier",
        "match_basis",
        "component_applicability",
        "exploitation_status",
        "kev_sources",
        "kev_date_added",
        "exploited_since",
        "evidence_confidence",
        "art14_readiness",
        "srp_readiness",
        "evidence_checked_at",
        "kev_snapshot_sha256",
        "euvd_data_status",
        "euvd_query_source",
        "euvd_last_successful_to_date",
        "euvd_reference_freshness",
        "euvd_source_db_sha256",
        "euvd_snapshot_created_at",
        "cra_review_required",
        "match_status",
        "confidence",
        "input_cve_ids",
        "input_euvd_ids",
        "component_name",
        "component_version",
        "component_vendor",
        "euvd_id",
        "severity",
        "cvss_score",
        "cvss_version",
        "epss_percent",
        "affected_product",
        "affected_vendor",
        "affected_versions",
        "alternative_ids",
        "published",
        "updated",
        "description",
        "match_reason",
        "references",
        "euvd_url",
        "component_purl",
        "component_cpe",
        "component_row",
    ]
    result_sheet = workbook.create_sheet("EUVD匹配结果")
    _write_rows(result_sheet, result_headers, results["matches"])
    _add_table(result_sheet, "EuvdMatches")
    if result_sheet.max_row >= 2:
        result_sheet.conditional_formatting.add(
            f"A2:A{result_sheet.max_row}",
            FormulaRule(formula=["$A2=\"EUVD精确匹配\""], fill=GREEN_FILL),
        )
        result_sheet.conditional_formatting.add(
            f"A2:A{result_sheet.max_row}",
            FormulaRule(formula=["$A2=\"产品候选匹配\""], fill=YELLOW_FILL),
        )

    review_rows = [row for row in results["matches"] if row["match_status"] == "需复核"]
    review_sheet = workbook.create_sheet("待人工确认")
    _write_rows(review_sheet, result_headers, review_rows)
    _add_table(review_sheet, "ReviewMatches")

    art14_rows = [row for row in results["matches"] if row.get("cra_review_required")]
    art14_sheet = workbook.create_sheet("Art14待评估")
    _write_rows(art14_sheet, result_headers, art14_rows)
    _add_table(art14_sheet, "Article14Review")

    component_headers = [
        "row_number",
        "name",
        "version",
        "vendor",
        "purl",
        "cpe",
        "scope",
        "license",
        "cve_ids",
        "euvd_ids",
        "confirmed_count",
        "review_count",
        "identity_ready",
        "query_status",
        "query_result_count",
        "query_api_total",
        "query_pages",
        "query_truncated",
        "query_mode",
        "unmapped_identifiers",
        "mapping_checked_at",
        "mapping_snapshot_sha256",
        "euvd_data_status",
        "euvd_query_source",
        "euvd_last_successful_to_date",
        "euvd_reference_freshness",
        "euvd_source_db_sha256",
        "euvd_snapshot_created_at",
        "result",
    ]
    component_sheet = workbook.create_sheet("组件清单")
    _write_rows(component_sheet, component_headers, results["components"])
    _add_table(component_sheet, "ComponentInventory")
    if component_sheet.max_row >= 2:
        result_column = get_column_letter(component_headers.index("result") + 1)
        component_sheet.conditional_formatting.add(
            f"{result_column}2:{result_column}{component_sheet.max_row}",
            FormulaRule(
                formula=[f'${result_column}2="发现EUVD记录"'],
                fill=RED_FILL,
            ),
        )
        component_sheet.conditional_formatting.add(
            f"{result_column}2:{result_column}{component_sheet.max_row}",
            FormulaRule(
                formula=[f'${result_column}2="需人工确认"'],
                fill=YELLOW_FILL,
            ),
        )

    input_snapshot = results.get("input_sbom_snapshot") or {}
    input_headers = list(input_snapshot.get("headers") or [])
    input_rows = list(input_snapshot.get("rows") or [])
    if input_headers:
        input_sheet = workbook.create_sheet("输入SBOM快照")
        _write_rows(input_sheet, input_headers, input_rows)
        _add_table(input_sheet, "InputSbomSnapshot")

    error_sheet = workbook.create_sheet("查询错误")
    _write_rows(error_sheet, ["component_row", "component_name", "error"], results["errors"])
    _add_table(error_sheet, "QueryErrors")

    notes_sheet = workbook.create_sheet("说明")
    notes_sheet.sheet_view.showGridLines = False
    notes = [
        ("项目", "EUVD Dependency Workbench"),
        ("数据范围", "报告使用 ENISA EUVD 的 CVE映射、详情和EU/CISA KEV公开快照；不查询SRP，不合并商业漏洞库。"),
        ("精确映射", "输入含CVE/EUVD ID时优先进行精确映射；没有漏洞标识符时才回退到产品名候选检索。"),
        ("EUVD记录匹配", "只确认标识符与公开EUVD记录的映射，不等于客户产品已确认受影响。"),
        ("KEV已知利用信号", "EU/CISA KEV列入表示存在公开已知利用情报；仍需进行产品适用性和产品级证据核验。"),
        ("未列入KEV", "只表示当前成功获取的KEV快照未列出，不代表漏洞未被利用。"),
        ("EPSS", "利用概率预测，不是已经发生恶意利用的证据。"),
        ("CRA Art.14", "自动化结果只能进入人工判定。依据CRA Art.3(42)与Art.14，需可靠证据、客户产品包含性及制造商awareness。"),
        ("SRP", "CRA Art.16 Single Reporting Platform是报告入口，不是公开查询数据库；本报告不代表已提交或已触发24h/72h时钟。"),
        ("需复核", "产品名匹配，但厂商缺失/差异较大，或 EUVD 版本表达式无法可靠机器解析。"),
        ("未发现匹配", "本次 EUVD 查询没有返回记录或候选；不是无漏洞声明，也不能排除SRP中存在非公开报告。"),
        ("身份覆盖率", "组件具备名称、版本，以及厂商/PURL/CPE 之一时计为身份信息可用于可靠匹配。"),
        ("检索覆盖率", "工具按 EUVD API 返回的总数逐页拉取；达到配置上限时会标记 query_truncated，不会静默当作完整结果。"),
        ("数据时效", "EUVD 在线结果会变化。报告中的生成时间用于审计和复现。"),
        ("官方入口", "https://euvd.enisa.europa.eu/"),
        ("官方 API", "https://euvd.enisa.europa.eu/apidoc"),
    ]
    notes_sheet.append(["主题", "说明"])
    _style_header(list(notes_sheet[1]))
    for row in notes:
        notes_sheet.append(row)
    notes_sheet.column_dimensions["A"].width = 18
    notes_sheet.column_dimensions["B"].width = 100
    for row in notes_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    notes_sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_template(output_path: Path) -> None:
    template_path = (
        Path(__file__).resolve().parent
        / "assets"
        / "客户SBOM导入模板_PRO-03B_v1.4兼容版.xlsx"
    )
    if template_path.is_file():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(template_path, output_path)
        return
    # The public source allowlist intentionally excludes the rights-pending
    # binary workbook.  Keep the endpoint functional with an independently
    # generated blank template whose source is part of this repository.
    write_public_template(output_path)
