from __future__ import annotations

import hashlib
import json
import tempfile
import zipfile
from calendar import monthrange
from datetime import datetime, timedelta, timezone
from html import escape
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .spreadsheet_io import _safe_cell


SRP_STAGES = {"early-warning", "notification", "final-report"}
CASE_TYPES = {"actively_exploited_vulnerability", "severe_incident"}
SRP_PROFILE_PATH = (
    Path(__file__).resolve().parent.parent / "config" / "srp-q16-2026-08-03.json"
)


def _load_srp_field_profile(path: Path = SRP_PROFILE_PATH) -> dict[str, Any]:
    """Load and fail closed on the versioned public ENISA Q16 field profile."""

    try:
        profile = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"无法加载 SRP 字段配置: {path}") from exc
    required = {
        "id",
        "faq_updated_at",
        "source_checked_at",
        "srp_information_url",
        "portal_url_status",
        "legend",
        "fields",
        "unnumbered_rows",
    }
    missing = sorted(required - set(profile))
    if missing:
        raise RuntimeError("SRP 字段配置缺少键: " + ", ".join(missing))
    fields = profile.get("fields")
    if not isinstance(fields, list) or not fields:
        raise RuntimeError("SRP 字段配置 fields 必须是非空数组")
    ids = [str(item.get("id") or "") for item in fields if isinstance(item, dict)]
    if len(ids) != len(fields) or any(not value for value in ids):
        raise RuntimeError("SRP 字段配置包含无效字段记录")
    if len(ids) != len(set(ids)):
        raise RuntimeError("SRP 字段配置包含重复字段 ID")
    expected_ids = {
        *(str(value) for value in range(1, 13)),
        *(f"v{value}" for value in range(13, 27)),
        *(f"i{value}" for value in range(13, 26)),
    }
    if set(ids) != expected_ids:
        raise RuntimeError("SRP 字段配置与受控 Q16 39 字段集合不一致")
    unnumbered_rows = profile.get("unnumbered_rows")
    if not isinstance(unnumbered_rows, list) or {
        str(item.get("id") or "") for item in unnumbered_rows
    } != {"i-final-description"}:
        raise RuntimeError("SRP 字段配置缺少 Q16 事件 Final 未编号描述行")
    for url_key in ("srp_information_url", "portal_url"):
        url = profile.get(url_key)
        if url is None and url_key == "portal_url":
            continue
        parsed = urlparse(str(url))
        host = (parsed.hostname or "").casefold()
        if parsed.scheme != "https" or not (
            host == "enisa.europa.eu" or host.endswith(".enisa.europa.eu")
        ):
            raise RuntimeError(f"SRP 字段配置 {url_key} 必须指向 ENISA HTTPS 域名")
    allowed_states = set(profile["legend"])
    for item in [*fields, *unnumbered_rows]:
        if item.get("scope") not in {"common", "vulnerability", "incident"}:
            raise RuntimeError(f"SRP 字段 {item['id']} scope 无效")
        for stage in SRP_STAGES:
            status = item.get(stage)
            if item.get("row_type") == "descriptor" and status is None:
                continue
            if status not in allowed_states:
                raise RuntimeError(f"SRP 字段 {item['id']} 的 {stage} 状态无效")
    return profile


SRP_FIELD_PROFILE = _load_srp_field_profile()


def _parse_aware(value: str | None) -> datetime | None:
    if not value:
        return None
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        raise ValueError("时间必须包含时区")
    return parsed.astimezone(timezone.utc)


def _case_type(case: dict[str, Any]) -> str:
    value = str(case.get("case_type") or "actively_exploited_vulnerability")
    if value not in CASE_TYPES:
        raise ValueError("Art.14 案件类型无效")
    return value


def _submission_time(case: dict[str, Any], stage: str) -> datetime | None:
    receipts = case.get("submission_receipts") or []
    timestamps = [
        _parse_aware(str(item.get("submitted_at") or ""))
        for item in receipts
        if item.get("stage") == stage and item.get("submitted_at")
    ]
    return min((value for value in timestamps if value is not None), default=None)


def _portal_value_map(case: dict[str, Any], stage: str) -> dict[str, Any]:
    fields = case.get("srp_fields") or {}
    case_type = _case_type(case)
    submission_times = {
        report_stage: _submission_time(case, report_stage)
        for report_stage in SRP_STAGES
    }
    return {
        "computed.notification_type": (
            "Severe Incident"
            if case_type == "severe_incident"
            else "Actively Exploited Vulnerability"
        ),
        "computed.notification_level": {
            "early-warning": "24h",
            "notification": "72h",
            "final-report": "Final",
        }[stage],
        "computed.reporting_time_24h": (
            submission_times["early-warning"].isoformat(timespec="seconds")
            if submission_times["early-warning"]
            else None
        ),
        "computed.reporting_time_72h": (
            submission_times["notification"].isoformat(timespec="seconds")
            if submission_times["notification"]
            else None
        ),
        "computed.reporting_time_final": (
            submission_times["final-report"].isoformat(timespec="seconds")
            if submission_times["final-report"]
            else None
        ),
        "computed.exploit_nature": fields.get("exploit_nature")
        or case.get("exploitation_evidence_summary"),
        "computed.corrective_measures_taken": fields.get(
            "corrective_measures_taken"
        )
        or case.get("mitigation_summary"),
        "computed.incident_corrective_measures_taken": fields.get(
            "incident_corrective_measures_taken"
        )
        or case.get("mitigation_summary"),
        **{f"case.{key}": value for key, value in case.items()},
        **{f"srp.{key}": value for key, value in fields.items()},
    }


def build_srp_portal_fields(case: dict[str, Any], stage: str) -> list[dict[str, Any]]:
    """Project local case data onto the exact ENISA FAQ Q16 stage matrix."""

    if stage not in SRP_STAGES:
        raise ValueError("SRP 报告阶段无效")
    case_type = _case_type(case)
    accepted_scopes = {
        "common",
        "incident" if case_type == "severe_incident" else "vulnerability",
    }
    values = _portal_value_map(case, stage)
    legend = SRP_FIELD_PROFILE["legend"]
    projected: list[dict[str, Any]] = []
    definitions = [
        *SRP_FIELD_PROFILE["fields"],
        *SRP_FIELD_PROFILE["unnumbered_rows"],
    ]
    for definition in definitions:
        if definition["scope"] not in accepted_scopes:
            continue
        source_status = definition[stage]
        is_descriptor = definition.get("row_type") == "descriptor"
        status = "D" if is_descriptor else source_status
        value = values.get(definition["value_key"])
        projected.append(
            {
                "id": definition["id"],
                "scope": definition["scope"],
                "label": definition["label"],
                "status": status,
                "status_meaning": (
                    "descriptive criterion row; Q16 assigns no stage code"
                    if is_descriptor
                    else legend[status]
                ),
                "value": value,
                "portal_automated": source_status == "A",
                "human_confirmation_required": source_status != "A",
                "q16_stage_code": source_status,
                "source": "ENISA CRA SRP FAQ Q16",
            }
        )
    return projected


def _add_calendar_month(value: datetime) -> datetime:
    year = value.year + (1 if value.month == 12 else 0)
    month = 1 if value.month == 12 else value.month + 1
    return value.replace(year=year, month=month, day=min(value.day, monthrange(year, month)[1]))


def deadline_status(case: dict[str, Any], now: datetime | None = None) -> dict[str, Any]:
    aware = _parse_aware(case.get("awareness_at"))
    corrective = _parse_aware(case.get("corrective_measure_available_at"))
    notification_submitted = _submission_time(case, "notification")
    case_type = _case_type(case)
    current = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)

    def item(due: datetime | None) -> dict[str, Any]:
        if due is None:
            return {"due_at": None, "status": "not_started", "remaining_seconds": None}
        remaining = int((due - current).total_seconds())
        return {
            "due_at": due.isoformat(timespec="seconds"),
            "status": "overdue" if remaining < 0 else "open",
            "remaining_seconds": remaining,
        }

    vulnerability_final = (
        corrective + timedelta(days=14)
        if case_type == "actively_exploited_vulnerability" and corrective
        else None
    )
    incident_final = (
        _add_calendar_month(notification_submitted)
        if case_type == "severe_incident" and notification_submitted
        else None
    )
    final_due = incident_final if case_type == "severe_incident" else vulnerability_final
    return {
        "basis": {
            "case_type": case_type,
            "awareness_is_manually_confirmed": bool(
                case.get("awareness_at") and case.get("awareness_confirmed_by")
            ),
            "automatic_awareness_inference": False,
            "reporting_rule": "without undue delay and no later than the calculated deadline",
            "final_report_basis": (
                "one calendar month after the 72h incident notification"
                if case_type == "severe_incident"
                else "14 days after a corrective or mitigating measure is available"
            ),
            "final_report_anchor_available": final_due is not None,
        },
        "early_warning_24h": item(aware + timedelta(hours=24) if aware else None),
        "notification_72h": item(aware + timedelta(hours=72) if aware else None),
        "final_report": item(final_due),
        "final_report_14d": item(vulnerability_final),
        "final_report_1m": item(incident_final),
    }


def srp_readiness(case: dict[str, Any], stage: str) -> dict[str, Any]:
    if stage not in SRP_STAGES:
        raise ValueError("SRP 报告阶段无效")
    fields = case.get("srp_fields") or {}
    case_type = _case_type(case)
    notification_type = (
        "Severe Incident"
        if case_type == "severe_incident"
        else "Actively Exploited Vulnerability"
    )
    common = {
        "notification_type": notification_type,
        "notification_level": stage,
        "reporter": fields.get("reporter"),
        "manufacturer_name": fields.get("manufacturer_name"),
        "product_name": case.get("project_name"),
        "product_type": fields.get("product_type"),
        "product_category": fields.get("product_category"),
        "member_states_where_available": fields.get("member_states_where_available"),
        "title": fields.get("title"),
        "cve_id": case.get("cve_id"),
        "euvd_id": case.get("euvd_id"),
    }
    stage_fields: dict[str, Any] = {}
    if case_type == "actively_exploited_vulnerability" and stage in {
        "notification",
        "final-report",
    }:
        stage_fields = {
            "general_information": fields.get("general_information"),
            "vulnerability_nature": fields.get("vulnerability_nature"),
            "exploit_nature": fields.get("exploit_nature")
            or case.get("exploitation_evidence_summary"),
            "corrective_measures_taken": fields.get("corrective_measures_taken")
            or case.get("mitigation_summary"),
            "user_measures": fields.get("user_measures"),
            "sensitivity": fields.get("sensitivity"),
        }
    if case_type == "actively_exploited_vulnerability" and stage == "final-report":
        stage_fields.update(
            {
                "corrective_measure_available_at": case.get(
                    "corrective_measure_available_at"
                ),
                "full_vulnerability_description": fields.get(
                    "full_vulnerability_description"
                ),
                "vulnerability_severity": fields.get("vulnerability_severity"),
                "vulnerability_impact": fields.get("vulnerability_impact"),
                "malicious_actor": fields.get("malicious_actor"),
                "security_update_details": fields.get("security_update_details"),
            }
        )
    if case_type == "severe_incident":
        stage_fields.update(
            {
                "suspected_unlawful_or_malicious_cause": fields.get(
                    "incident_suspected_unlawful_or_malicious"
                ),
            }
        )
        if stage in {"notification", "final-report"}:
            stage_fields.update(
                {
                    "incident_general_nature": fields.get("incident_general_nature"),
                    "incident_detected_at": fields.get("incident_detected_at"),
                    "incident_occurred_at": fields.get("incident_occurred_at"),
                    "incident_initial_assessment": fields.get(
                        "incident_initial_assessment"
                    ),
                    "incident_corrective_measures_taken": fields.get(
                        "incident_corrective_measures_taken"
                    )
                    or case.get("mitigation_summary"),
                    "incident_user_measures": fields.get("incident_user_measures"),
                    "sensitivity": fields.get("sensitivity"),
                }
            )
        if stage == "final-report":
            stage_fields.update(
                {
                    "incident_detailed_description": fields.get(
                        "incident_detailed_description"
                    ),
                    "incident_severity": fields.get("incident_severity"),
                    "incident_impact": fields.get("incident_impact"),
                    "incident_likely_threat_or_root_cause": fields.get(
                        "incident_likely_threat_or_root_cause"
                    ),
                    "incident_applied_and_ongoing_mitigation_measures": fields.get(
                        "incident_applied_and_ongoing_mitigation_measures"
                    ),
                }
            )
    payload = {**common, **stage_fields}
    required = {
        "manufacturer_name",
        "product_name",
        "title",
    }
    if case_type == "actively_exploited_vulnerability" and stage in {
        "notification",
        "final-report",
    }:
        required.update(
            {
                "general_information",
                "vulnerability_nature",
                "exploit_nature",
                "corrective_measures_taken",
                "user_measures",
            }
        )
    if case_type == "actively_exploited_vulnerability" and stage == "final-report":
        required.update(
            {
                "corrective_measure_available_at",
                "full_vulnerability_description",
                "vulnerability_severity",
                "vulnerability_impact",
                "security_update_details",
            }
        )
    if case_type == "severe_incident":
        required.update(
            {
                "suspected_unlawful_or_malicious_cause",
            }
        )
        if stage in {"notification", "final-report"}:
            required.update(
                {
                    "incident_general_nature",
                    "incident_detected_at",
                    "incident_occurred_at",
                    "incident_initial_assessment",
                    "incident_corrective_measures_taken",
                    "incident_user_measures",
                }
            )
        if stage == "final-report":
            required.update(
                {
                    "incident_detailed_description",
                    "incident_severity",
                    "incident_impact",
                    "incident_likely_threat_or_root_cause",
                    "incident_applied_and_ongoing_mitigation_measures",
                }
            )
    if fields.get("product_type") and str(fields["product_type"]).casefold() != "default":
        required.add("product_category")
    missing = [key for key in sorted(required) if not payload.get(key)]
    severe_criteria = case.get("severe_incident_criteria") or {}
    prerequisite_stages = {
        "early-warning": [],
        "notification": ["early-warning"],
        "final-report": ["early-warning", "notification"],
    }[stage]
    submitted_stages = {
        str(item.get("stage")) for item in case.get("submission_receipts") or []
    }
    missing_prerequisite_receipts = [
        item for item in prerequisite_stages if item not in submitted_stages
    ]
    gates = {
        "four_eye_approved": case.get("workflow_status")
        in {"approved", "submitted"},
        "decision_reportable": case.get("art14_decision") == "reportable",
        "awareness_confirmed": bool(
            case.get("awareness_at") and case.get("awareness_confirmed_by")
        ),
        "manual_submission_only": True,
        "srp_api_submission": False,
        "severe_incident_criteria_met": (
            bool(
                severe_criteria.get(
                    "availability_authenticity_integrity_confidentiality_impact"
                )
                or severe_criteria.get("malicious_code_introduction")
            )
            if case_type == "severe_incident"
            else True
        ),
        "severe_incident_criteria_rationale_recorded": (
            bool(str(severe_criteria.get("rationale") or "").strip())
            if case_type == "severe_incident"
            else True
        ),
        "portal_field_confirmation_required": True,
        "previous_stage_receipts_recorded": not missing_prerequisite_receipts,
    }
    ready = not missing and all(
        gates[key]
        for key in (
            "four_eye_approved",
            "decision_reportable",
            "awareness_confirmed",
            "severe_incident_criteria_met",
            "severe_incident_criteria_rationale_recorded",
        )
    )
    return {
        "stage": stage,
        "case_type": case_type,
        "schema_profile": SRP_FIELD_PROFILE,
        "ready": ready,
        "material_ready": ready,
        "portal_submission_ready": ready and not missing_prerequisite_receipts,
        "missing_prerequisite_receipts": missing_prerequisite_receipts,
        "missing_fields": missing,
        "conditional_fields": {
            "member_states_where_available": "obligatory_if_information_available",
            "sensitivity": "obligatory_if_information_available",
            "product_category": "applies_if_product_type_is_not_default",
            **(
                {"malicious_actor": "obligatory_if_information_available"}
                if case_type == "actively_exploited_vulnerability"
                and stage == "final-report"
                else {}
            ),
        },
        "gates": gates,
    }


def build_srp_payload(case: dict[str, Any], stage: str) -> dict[str, Any]:
    readiness = srp_readiness(case, stage)
    fields = case.get("srp_fields") or {}
    case_type = _case_type(case)
    notification_type = (
        "Severe Incident"
        if case_type == "severe_incident"
        else "Actively Exploited Vulnerability"
    )
    submission_times = {
        report_stage: _submission_time(case, report_stage)
        for report_stage in ("early-warning", "notification", "final-report")
    }
    payload = {
        "document_type": "CRA Article 14 SRP preparation draft",
        "schema_profile": SRP_FIELD_PROFILE,
        "case_type": case_type,
        "stage": stage,
        "draft_only": True,
        "automatic_submission": False,
        "submission_mode": "manual_only",
        "assistance_mode": "generate_review_confirm_open_official_portal",
        "official_submission_performed": False,
        "official_submission_receipt": None,
        "legal_basis": (
            ["CRA Art.14(3)-(7)", "CRA Art.16"]
            if case_type == "severe_incident"
            else ["CRA Art.3(42)", "CRA Art.14(1)-(2),(7)", "CRA Art.16"]
        ),
        "case_id": case["id"],
        "common_fields": {
            "notification_type": notification_type,
            "notification_level": stage,
            "reporter": fields.get("reporter"),
            "reporting_times": {
                "early_warning_24h": (
                    submission_times["early-warning"].isoformat(timespec="seconds")
                    if submission_times["early-warning"]
                    else None
                ),
                "notification_72h": (
                    submission_times["notification"].isoformat(timespec="seconds")
                    if submission_times["notification"]
                    else None
                ),
                "final_report": (
                    submission_times["final-report"].isoformat(timespec="seconds")
                    if submission_times["final-report"]
                    else None
                ),
            },
            "title": fields.get("title"),
        },
        "manufacturer": {
            "name": fields.get("manufacturer_name"),
            "contact": fields.get("manufacturer_contact"),
        },
        "product": {
            "name": case.get("project_name"),
            "version": case.get("project_version"),
            "build": case.get("software_build"),
            "type": fields.get("product_type"),
            "category": fields.get("product_category"),
            "member_states_where_available": fields.get(
                "member_states_where_available"
            ),
            "component": case.get("component_name"),
            "component_version": case.get("component_version"),
        },
        "vulnerability": {
            "cve_id": case.get("cve_id"),
            "euvd_id": case.get("euvd_id"),
            "product_applicability": case.get("applicability_status"),
            "applicability_justification": case.get("applicability_justification"),
        }
        if case_type == "actively_exploited_vulnerability"
        else None,
        "exploitation": {
            "evidence_status": case.get("exploitation_evidence_status"),
            "nature": fields.get("exploit_nature")
            or case.get("exploitation_evidence_summary"),
            "external_signal_at": case.get("external_signal_at"),
            "manufacturer_awareness_at": case.get("awareness_at"),
            "awareness_confirmed_by": case.get("awareness_confirmed_by"),
        }
        if case_type == "actively_exploited_vulnerability"
        else None,
        "incident": {
            "severe_incident_criteria": case.get("severe_incident_criteria") or {},
            "suspected_unlawful_or_malicious_cause": fields.get(
                "incident_suspected_unlawful_or_malicious"
            ),
            "general_nature": fields.get("incident_general_nature"),
            "detected_at": fields.get("incident_detected_at"),
            "occurred_at": fields.get("incident_occurred_at"),
            "initial_assessment": fields.get("incident_initial_assessment"),
            "corrective_measures_taken": fields.get(
                "incident_corrective_measures_taken"
            )
            or case.get("mitigation_summary"),
            "user_measures": fields.get("incident_user_measures"),
            "detailed_description": fields.get("incident_detailed_description"),
            "severity": fields.get("incident_severity"),
            "impact": fields.get("incident_impact"),
            "likely_threat_or_root_cause": fields.get(
                "incident_likely_threat_or_root_cause"
            ),
            "applied_and_ongoing_mitigation_measures": fields.get(
                "incident_applied_and_ongoing_mitigation_measures"
            ),
        }
        if case_type == "severe_incident"
        else None,
        "risk_and_response": {
            "product_risk_summary": case.get("product_risk_summary"),
            "general_information": fields.get("general_information"),
            "vulnerability_nature": fields.get("vulnerability_nature"),
            "corrective_measures_taken": fields.get("corrective_measures_taken")
            or case.get("mitigation_summary"),
            "user_measures": fields.get("user_measures"),
            "csirt_coordinator": fields.get("csirt_coordinator"),
            "user_notification": fields.get("user_notification"),
            "sensitivity": fields.get("sensitivity"),
        },
        "final_report": {
            "corrective_measure_available_at": case.get(
                "corrective_measure_available_at"
            ),
            "full_vulnerability_description": fields.get(
                "full_vulnerability_description"
            ),
            "vulnerability_severity": fields.get("vulnerability_severity"),
            "vulnerability_impact": fields.get("vulnerability_impact"),
            "malicious_actor": fields.get("malicious_actor"),
            "security_update_details": fields.get("security_update_details"),
            "remediation_monitoring": fields.get("remediation_monitoring"),
        }
        if stage == "final-report" and case_type == "actively_exploited_vulnerability"
        else None,
        "approval": {
            "technical_reviewer": case.get("technical_reviewer"),
            "technical_reviewed_at": case.get("technical_reviewed_at"),
            "compliance_reviewer": case.get("compliance_reviewer"),
            "compliance_reviewed_at": case.get("compliance_reviewed_at"),
            "decision": case.get("art14_decision"),
        },
        "local_supporting_context": {
            "manufacturer_contact": fields.get("manufacturer_contact"),
            "csirt_coordinator": fields.get("csirt_coordinator"),
            "user_notification": fields.get("user_notification"),
            "remediation_monitoring": fields.get("remediation_monitoring"),
            "note": "Local workflow context; not represented as an ENISA Q16 portal field.",
        },
        "portal_fields": build_srp_portal_fields(case, stage),
        "human_confirmation": {
            "required": True,
            "confirmed": False,
            "confirmation_scope": (
                "Compare every generated value with the live ENISA SRP form, "
                "confirm the notification type, stage, deadline, sensitivity, "
                "Member States and any copied values, then submit in the official portal."
            ),
            "tool_submit_button_is_not_provided": True,
        },
        "deadlines": deadline_status(case),
        "readiness": readiness,
    }
    return payload


def write_srp_json(path: Path, case: dict[str, Any], stage: str) -> None:
    path.write_text(
        json.dumps(build_srp_payload(case, stage), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_srp_xlsx(path: Path, case: dict[str, Any], stage: str) -> None:
    payload = build_srp_payload(case, stage)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SRP草稿"
    sheet.append(["字段", "内容"])
    header_fill = PatternFill("solid", fgColor="163B65")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    rows = [
        ("声明", "仅供人工审核和在 ENISA SRP 手工填报；本文件不会自动提交。"),
        ("字段配置版本", payload["schema_profile"]["id"]),
        ("案件类型", payload["case_type"]),
        ("阶段", stage),
        ("通知类型", payload["common_fields"]["notification_type"]),
        ("通知级别", payload["common_fields"]["notification_level"]),
        ("Reporter", payload["common_fields"]["reporter"]),
        ("标题", payload["common_fields"]["title"]),
        ("案件ID", case["id"]),
        ("制造商", payload["manufacturer"]["name"]),
        ("制造商联系方式", payload["manufacturer"]["contact"]),
        ("产品", payload["product"]["name"]),
        ("产品版本", payload["product"]["version"]),
        ("构建", payload["product"]["build"]),
        ("组件", payload["product"]["component"]),
        ("awareness", case.get("awareness_at")),
        ("风险分析", payload["risk_and_response"]["product_risk_summary"]),
        ("产品可用成员国", payload["product"]["member_states_where_available"]),
        ("CSIRT协调员", payload["risk_and_response"]["csirt_coordinator"]),
        ("用户通知", payload["risk_and_response"]["user_notification"]),
        ("敏感性", payload["risk_and_response"]["sensitivity"]),
        ("技术审核人", payload["approval"]["technical_reviewer"]),
        ("合规审核人", payload["approval"]["compliance_reviewer"]),
        ("Art.14决定", payload["approval"]["decision"]),
        ("草稿准备状态", "Ready" if payload["readiness"]["ready"] else "Not ready"),
        ("缺少字段", ", ".join(payload["readiness"]["missing_fields"])),
    ]
    if payload["case_type"] == "actively_exploited_vulnerability":
        vulnerability = payload["vulnerability"] or {}
        exploitation = payload["exploitation"] or {}
        rows.extend(
            [
                ("CVE", vulnerability.get("cve_id")),
                ("EUVD", vulnerability.get("euvd_id")),
                ("产品适用性", vulnerability.get("product_applicability")),
                ("适用性理由", vulnerability.get("applicability_justification")),
                ("积极利用证据", exploitation.get("evidence_status")),
                ("利用性质", exploitation.get("nature")),
                ("一般信息", payload["risk_and_response"]["general_information"]),
                ("漏洞一般性质", payload["risk_and_response"]["vulnerability_nature"]),
                ("已采取修正/缓解措施", payload["risk_and_response"]["corrective_measures_taken"]),
                ("用户可采取措施", payload["risk_and_response"]["user_measures"]),
            ]
        )
        if stage == "final-report":
            rows.extend(
                [
                    ("修正措施可用时间", (payload["final_report"] or {}).get("corrective_measure_available_at")),
                    ("完整漏洞说明", (payload["final_report"] or {}).get("full_vulnerability_description")),
                    ("漏洞严重性", (payload["final_report"] or {}).get("vulnerability_severity")),
                    ("漏洞影响", (payload["final_report"] or {}).get("vulnerability_impact")),
                    ("恶意行为者", (payload["final_report"] or {}).get("malicious_actor")),
                    ("安全更新详情", (payload["final_report"] or {}).get("security_update_details")),
                    ("修复后监测", (payload["final_report"] or {}).get("remediation_monitoring")),
                ]
            )
    else:
        incident = payload["incident"] or {}
        rows.extend(
            [
                ("Art.14(5) 严重事件准则", json.dumps(incident.get("severe_incident_criteria") or {}, ensure_ascii=False)),
                ("疑似非法/恶意原因", incident.get("suspected_unlawful_or_malicious_cause")),
                ("事件一般性质", incident.get("general_nature")),
                ("事件检测时间", incident.get("detected_at")),
                ("事件发生时间", incident.get("occurred_at")),
                ("初步评估", incident.get("initial_assessment")),
                ("已采取修正措施", incident.get("corrective_measures_taken")),
                ("用户可采取措施", incident.get("user_measures")),
            ]
        )
        if stage == "final-report":
            rows.extend(
                [
                    ("事件详细说明", incident.get("detailed_description")),
                    ("事件严重性", incident.get("severity")),
                    ("事件影响", incident.get("impact")),
                    ("可能威胁/根因", incident.get("likely_threat_or_root_cause")),
                    ("已应用及进行中的缓解措施", incident.get("applied_and_ongoing_mitigation_measures")),
                ]
            )
    # All values can ultimately originate from an uploaded SBOM or an analyst
    # form.  Route both columns through the same formula/XML guard used by the
    # main report writer so Excel never interprets customer text as a formula.
    for key, value in rows:
        sheet.append([_safe_cell(key), _safe_cell(value)])
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 90
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    portal_sheet = workbook.create_sheet("Q16门户核对")
    portal_sheet.append(["确认", "Q16 ID", "字段", "本阶段状态", "生成值"])
    for cell in portal_sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for item in payload["portal_fields"]:
        portal_sheet.append(
            [
                _safe_cell("门户自动填充" if item["portal_automated"] else "待人工确认"),
                _safe_cell(item["id"]),
                _safe_cell(item["label"]),
                _safe_cell(
                    f"{item['status']} — {item['status_meaning']}"
                ),
                _safe_cell(
                    json.dumps(item["value"], ensure_ascii=False)
                    if isinstance(item["value"], (dict, list))
                    else item["value"]
                ),
            ]
        )
    portal_sheet.column_dimensions["A"].width = 16
    portal_sheet.column_dimensions["B"].width = 12
    portal_sheet.column_dimensions["C"].width = 55
    portal_sheet.column_dimensions["D"].width = 45
    portal_sheet.column_dimensions["E"].width = 90
    for row in portal_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def write_srp_html(path: Path, case: dict[str, Any], stage: str) -> None:
    payload = build_srp_payload(case, stage)
    rows = []
    for section, value in payload.items():
        if isinstance(value, dict):
            for key, item in value.items():
                rows.append((f"{section}.{key}", json.dumps(item, ensure_ascii=False) if isinstance(item, (dict, list)) else item))
        else:
            rows.append(
                (
                    section,
                    json.dumps(value, ensure_ascii=False)
                    if isinstance(value, list)
                    else value,
                )
            )
    body = "\n".join(
        f"<tr><th>{escape(str(key))}</th><td>{escape('' if value is None else str(value))}</td></tr>"
        for key, value in rows
    )
    path.write_text(
        f"""<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"><title>SRP {escape(stage)} draft</title>
<style>body{{font-family:Arial,sans-serif;margin:32px;color:#17202a}}h1{{font-size:24px}}
.notice{{padding:12px;background:#fff4ce;border:1px solid #e0b000}}table{{border-collapse:collapse;width:100%;margin-top:18px}}
th,td{{border:1px solid #ccd4dc;padding:8px;text-align:left;vertical-align:top}}th{{width:28%;background:#eef4f8}}
@media print{{.notice{{break-inside:avoid}}}}</style></head><body>
<h1>CRA Article 14 SRP {escape(stage)} 草稿</h1>
<p class="notice">仅供人工审核和手工填报。未连接或自动提交至 ENISA SRP。</p>
<table>{body}</table></body></html>""",
        encoding="utf-8",
    )


def _markdown_cell(value: Any) -> str:
    if value is None or value == "":
        return "—"
    rendered = (
        json.dumps(value, ensure_ascii=False, sort_keys=True)
        if isinstance(value, (dict, list))
        else str(value)
    )
    return (
        escape(rendered)
        .replace("|", "\\|")
        .replace("\r", " ")
        .replace("\n", "<br>")
    )


def _portal_field_checklist(case: dict[str, Any], stage: str) -> str:
    payload = build_srp_payload(case, stage)
    lines = [
        "# ENISA SRP 门户字段逐项核对表",
        "",
        f"- 案件 ID：`{case['id']}`",
        f"- 阶段：`{stage}`",
        f"- 字段配置：`{SRP_FIELD_PROFILE['id']}`",
        f"- ENISA FAQ 日期：`{SRP_FIELD_PROFILE['faq_updated_at']}`",
        "- 状态：本地准备草稿；不是官方提交、回执或法律结论。",
        "",
        "| 确认 | Q16 ID | 门户字段 | 本阶段状态 | 生成值 |",
        "|---|---|---|---|---|",
    ]
    for item in payload["portal_fields"]:
        check = "门户自动填充" if item["portal_automated"] else "☐"
        lines.append(
            "| "
            + " | ".join(
                [
                    check,
                    str(item["id"]),
                    _markdown_cell(item["label"]),
                    f"{item['status']} — {_markdown_cell(item['status_meaning'])}",
                    _markdown_cell(item["value"]),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 提交前人工确认",
            "",
            "- [ ] 已在当前 ENISA SRP 页面核对字段名称、必填性和帮助文本。",
            "- [ ] 已确认案件确属积极利用漏洞或 CRA Art.14(5) 严重事件。",
            "- [ ] 已确认制造商 awareness 起算点以及 24h/72h/Final 最迟期限。",
            "- [ ] 已确认制造商、产品、产品类型/类别和成员国范围。",
            "- [ ] 已确认敏感性、PEC/延迟分发需求及用户通知安排。",
            "- [ ] 已确认本包没有把 EUVD/KEV 信号自动当成 Art.14 结论。",
            "- [ ] 已由授权人员在官方 SRP 中点击 Submit，并保存通知 ID、状态、时间及邮件/告警。",
            "",
        ]
    )
    return "\n".join(lines)


def _human_submission_guide(case: dict[str, Any], stage: str) -> str:
    readiness = srp_readiness(case, stage)
    previous = readiness["missing_prerequisite_receipts"]
    sequence_note = (
        "前序官方回执已齐全。"
        if not previous
        else "尚缺前序官方回执：" + ", ".join(previous) + "。可提前准备材料，但不可提交本阶段。"
    )
    return f"""# 人工确认与官方提交说明

本包由 EUVD Dependency Workbench 为案件 `{case['id']}` 的 `{stage}` 阶段生成。

## 结论边界

- 本包只准备字段与复核材料，不连接 ENISA SRP，不保存 EU Login 凭据，也不会点击官方 Submit。
- `ready={str(readiness['ready']).lower()}` 仅表示本地必填字段和工作流门已满足，不代表 ENISA 接受或法定期限已履行。
- {sequence_note}
- 官方提交成立的唯一证据，是 SRP 返回的通知 ID/状态、提交时间以及相应邮件或告警。

## 操作顺序

1. 打开 `PORTAL_FIELD_CHECKLIST.md`，逐项与当前 SRP 表单核对。
2. 使用处于 Active 状态的 AR 账户登录官方 SRP，选择正确的制造商与 CDaC。
3. 按 `{stage}` 阶段录入或更新字段；需要时先保存 Draft。
4. 由授权人员进行最终人工确认后，在官方 SRP 点击 Submit。
5. 保存通知 ID、门户状态、提交时间、确认邮件/告警；回到 Workbench 登记手工提交回执。
6. Final Report 提交后门户记录不可编辑；提交前须再次核对完整性。

## 当前入口

ENISA SRP 信息页：{SRP_FIELD_PROFILE['srp_information_url']}

正式门户 URL 状态：`{SRP_FIELD_PROFILE['portal_url_status']}`。如 ENISA 已发布新入口，以实时官方页面为准。
"""


def write_srp_submission_package_zip(
    path: Path,
    case: dict[str, Any],
    stage: str,
) -> dict[str, Any]:
    """Write a reviewable, hash-bound SRP preparation package for one stage."""

    readiness = srp_readiness(case, stage)
    if not readiness["ready"]:
        detail = ", ".join(readiness["missing_fields"]) or "审批门未满足"
        raise ValueError("SRP 完整上报包尚未就绪: " + detail)
    if path.is_symlink():
        raise ValueError("拒绝写入符号链接形式的 SRP 上报包目标")
    with tempfile.TemporaryDirectory(prefix="euvd-srp-package-") as directory:
        working = Path(directory)
        material_paths = {
            "SRP_DRAFT.json": working / "SRP_DRAFT.json",
            "SRP_DRAFT.xlsx": working / "SRP_DRAFT.xlsx",
            "SRP_DRAFT.html": working / "SRP_DRAFT.html",
        }
        write_srp_json(material_paths["SRP_DRAFT.json"], case, stage)
        write_srp_xlsx(material_paths["SRP_DRAFT.xlsx"], case, stage)
        write_srp_html(material_paths["SRP_DRAFT.html"], case, stage)

        materials: dict[str, bytes] = {
            name: source.read_bytes() for name, source in material_paths.items()
        }
        materials["PORTAL_FIELD_CHECKLIST.md"] = _portal_field_checklist(
            case, stage
        ).encode("utf-8")
        materials["HUMAN_REVIEW_AND_SUBMISSION.md"] = _human_submission_guide(
            case, stage
        ).encode("utf-8")
        evidence_index = {
            "case_id": case["id"],
            "stage": stage,
            "source_evidence_included": False,
            "hash_scope": (
                "files lists all submission materials; SHA256SUMS additionally "
                "binds PACKAGE_MANIFEST.json and excludes only SHA256SUMS itself"
            ),
            "note": (
                "This index carries evidence metadata only. Source evidence remains "
                "under the manufacturer's access control unless the live portal requests it."
            ),
            "evidence": case.get("evidence") or [],
            "approvals": case.get("approvals") or [],
        }
        materials["EVIDENCE_INDEX.json"] = (
            json.dumps(evidence_index, ensure_ascii=False, indent=2) + "\n"
        ).encode("utf-8")

        hashes = {
            name: hashlib.sha256(content).hexdigest()
            for name, content in sorted(materials.items())
        }
        generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
        manifest = {
            "package_type": "CRA_ARTICLE_14_SRP_ASSISTED_SUBMISSION_PACKAGE",
            "case_id": case["id"],
            "case_type": _case_type(case),
            "stage": stage,
            "generated_at": generated_at,
            "schema_profile_id": SRP_FIELD_PROFILE["id"],
            "readiness": readiness,
            "human_confirmation_required": True,
            "automatic_submission": False,
            "official_submission_performed": False,
            "official_submission_receipt": None,
            "source_evidence_included": False,
            "files": [
                {"path": name, "sha256": digest, "size": len(materials[name])}
                for name, digest in hashes.items()
            ],
        }
        materials["PACKAGE_MANIFEST.json"] = (
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n"
        ).encode("utf-8")
        sums_hashes = {
            name: hashlib.sha256(content).hexdigest()
            for name, content in sorted(materials.items())
        }
        materials["SHA256SUMS"] = (
            "\n".join(
                f"{digest}  {name}" for name, digest in sums_hashes.items()
            )
            + "\n"
        ).encode("utf-8")

        path.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(
            path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
        ) as archive:
            for name, content in materials.items():
                archive.writestr(name, content)
    return manifest
