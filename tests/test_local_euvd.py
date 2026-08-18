from __future__ import annotations

import asyncio
import hashlib
import json
import sqlite3
import tempfile
import unittest
from contextlib import closing
from pathlib import Path

import httpx
from openpyxl import load_workbook

from app import matcher
from app.spreadsheet_io import write_report


class LocalEuvdSnapshotTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.database_path = Path(self.temporary.name) / "euvd-readonly.sqlite3"
        record = {
            "id": "EUVD-2024-1",
            "aliases": "CVE-2024-1000",
            "enisaIdProduct": [
                {
                    "product": {"name": "Example Product", "vendor": {"name": "Example"}},
                    "product_version": "1.0",
                }
            ],
        }
        with closing(sqlite3.connect(self.database_path)) as connection, connection:
            connection.executescript(
                """
                CREATE TABLE web_snapshot_metadata(key TEXT PRIMARY KEY, value TEXT NOT NULL);
                CREATE TABLE data_source_versions(
                    source_name TEXT, last_checked_at TEXT, last_modified TEXT,
                    content_sha256 TEXT, record_count INTEGER, http_status INTEGER,
                    source_url TEXT, is_current INTEGER
                );
                CREATE TABLE cve_euvd_mapping(cve_id TEXT, euvd_id TEXT);
                CREATE TABLE known_exploited(cve_id TEXT, euvd_id TEXT, date_added TEXT);
                CREATE TABLE known_exploited_sources(cve_id TEXT, euvd_id TEXT, source TEXT);
                CREATE TABLE vulnerabilities(euvd_id TEXT PRIMARY KEY, record_json TEXT, updated_raw TEXT);
                CREATE TABLE local_products(
                    product_key TEXT, vendor_key TEXT, euvd_id TEXT,
                    product_name TEXT, vendor_name TEXT, affected_versions TEXT
                );
                """
            )
            metadata = {
                "source_id": "EUVD_LOCAL_MIRROR",
                "snapshot_created_at": "2026-08-04T00:00:00+00:00",
                "last_successful_to_date": "2026-07-28",
                "reference_data_freshness": "degraded_local_snapshot",
                "vulnerability_count": "1",
                "mapping_count": "1",
                "known_exploited_count": "1",
                "product_index_count": "1",
                "source_db_sha256": "a" * 64,
                "snapshot_integrity_check_at_build": "ok",
                "consumer_boundary": "read-only consumer",
                "current_source_versions_json": json.dumps(
                    [
                        {
                            "source_name": "cve_euvd_mapping",
                            "last_modified": "2026-07-30T00:00:00+00:00",
                            "content_sha256": "b" * 64,
                            "record_count": 1,
                            "http_status": 0,
                        }
                    ]
                ),
            }
            connection.executemany(
                "INSERT INTO web_snapshot_metadata VALUES (?, ?)", metadata.items()
            )
            for name, digest in (("cve_euvd_mapping", "b" * 64), ("known_exploited", "c" * 64)):
                connection.execute(
                    "INSERT INTO data_source_versions VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (name, "2026-08-04T00:00:00Z", "2026-07-30T00:00:00Z", digest, 1, 0, "local-snapshot://test", 1),
                )
            connection.execute(
                "INSERT INTO cve_euvd_mapping VALUES (?, ?)",
                ("CVE-2024-1000", "EUVD-2024-1"),
            )
            connection.execute(
                "INSERT INTO known_exploited VALUES (?, ?, ?)",
                ("CVE-2024-1000", "EUVD-2024-1", "2026-01-01"),
            )
            connection.execute(
                "INSERT INTO known_exploited_sources VALUES (?, ?, ?)",
                ("CVE-2024-1000", "EUVD-2024-1", "cisa_kev"),
            )
            connection.execute(
                "INSERT INTO vulnerabilities VALUES (?, ?, ?)",
                ("EUVD-2024-1", json.dumps(record), "2026-01-01"),
            )
            connection.execute(
                "INSERT INTO local_products VALUES (?, ?, ?, ?, ?, ?)",
                (
                    "exampleproduct",
                    "example",
                    "EUVD-2024-1",
                    "Example Product",
                    "Example",
                    "1.0",
                ),
            )
        self.previous_path = matcher.LOCAL_EUVD_DB
        self.previous_sha_path = matcher.LOCAL_EUVD_SHA256_FILE
        self.previous_expected_sha = matcher.LOCAL_EUVD_EXPECTED_SHA256
        self.previous_network_fallback = matcher.NETWORK_FALLBACK
        matcher.LOCAL_EUVD_DB = self.database_path
        matcher.LOCAL_EUVD_SHA256_FILE = Path(
            str(self.database_path) + ".sha256"
        )
        matcher.LOCAL_EUVD_SHA256_FILE.write_text(
            hashlib.sha256(self.database_path.read_bytes()).hexdigest() + "\n",
            encoding="ascii",
        )
        matcher.LOCAL_EUVD_EXPECTED_SHA256 = ""
        matcher.NETWORK_FALLBACK = False
        matcher.EuvdClient._local_validation_cache = None

    def tearDown(self) -> None:
        matcher.LOCAL_EUVD_DB = self.previous_path
        matcher.LOCAL_EUVD_SHA256_FILE = self.previous_sha_path
        matcher.LOCAL_EUVD_EXPECTED_SHA256 = self.previous_expected_sha
        matcher.NETWORK_FALLBACK = self.previous_network_fallback
        matcher.EuvdClient._local_validation_cache = None
        self.temporary.cleanup()

    def test_status_preserves_degraded_freshness_boundary(self) -> None:
        status = matcher.EuvdClient.local_snapshot_status()
        self.assertEqual(status["status"], "local_degraded")
        self.assertEqual(status["integrity_check_at_build"], "ok")
        self.assertFalse(status["network_required_for_queries"])

    def test_mapping_kev_detail_and_product_query_are_local(self) -> None:
        async def exercise() -> None:
            wrapper = matcher.EuvdClient()
            async with httpx.AsyncClient() as client:
                mapping, mapping_meta = await wrapper.cve_mapping(
                    client, {"CVE-2024-1000"}
                )
                kev, kev_meta = await wrapper.kev_index(client)
                detail = await wrapper.detail(client, "EUVD-2024-1")
                items, query_meta = await wrapper.search(
                    client, "Example Product", "Example"
                )
            result = await matcher.match_components(
                [
                    matcher.Component(
                        row_number=2,
                        name="Example Product",
                        version="1.0",
                        vendor="Example",
                    )
                ]
            )
            self.assertEqual(mapping["CVE-2024-1000"], ["EUVD-2024-1"])
            self.assertEqual(mapping_meta["mapping_snapshot_source"], "local-read-only-mirror")
            self.assertEqual(kev["CVE-2024-1000"]["sources"], ["cisa_kev"])
            self.assertEqual(kev_meta["kev_freshness"], "degraded_local_snapshot")
            self.assertEqual(detail["id"], "EUVD-2024-1")
            self.assertEqual([item["id"] for item in items], ["EUVD-2024-1"])
            self.assertEqual(query_meta["query_mode"], "local-mirror-product-exact")
            self.assertEqual(result["data_provenance"]["status"], "local_degraded")
            self.assertEqual(
                result["components"][0]["euvd_last_successful_to_date"],
                "2026-07-28",
            )
            self.assertEqual(
                result["matches"][0]["euvd_reference_freshness"],
                "degraded_local_snapshot",
            )
            report_path = Path(self.temporary.name) / "local-euvd-report.xlsx"
            write_report(report_path, "test.cdx.json", result)
            workbook = load_workbook(report_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook["汇总"]["B9"].value, "2026-07-28")
                headers = [cell.value for cell in next(workbook["组件清单"].iter_rows())]
                cutoff_column = headers.index("euvd_last_successful_to_date") + 1
                self.assertEqual(
                    workbook["组件清单"].cell(2, cutoff_column).value,
                    "2026-07-28",
                )
            finally:
                workbook.close()

        asyncio.run(exercise())

    def _add_second_product(
        self, vendor_name: str, euvd_id: str, product_version: str = "2.0"
    ) -> None:
        """Insert a second record sharing the product_key but differing vendor,
        so vendor+product joint matching has something to narrow against."""
        record = {
            "id": euvd_id,
            "aliases": "CVE-2024-2000",
            "enisaIdProduct": [
                {
                    "product": {"name": "Example Product", "vendor": {"name": vendor_name}},
                    "product_version": product_version,
                }
            ],
        }
        product_key = "exampleproduct"
        vendor_key = matcher.normalize_key(vendor_name)
        with closing(sqlite3.connect(self.database_path)) as connection:
            connection.execute(
                "INSERT INTO vulnerabilities VALUES (?, ?, ?)",
                (euvd_id, json.dumps(record), "2026-01-02"),
            )
            connection.execute(
                "INSERT INTO local_products VALUES (?, ?, ?, ?, ?, ?)",
                (
                    product_key,
                    vendor_key,
                    euvd_id,
                    "Example Product",
                    vendor_name,
                    product_version,
                ),
            )
            connection.commit()
        # The DB bytes changed; recompute the sidecar sha and drop the client's
        # validation cache so the next search passes the integrity gate.
        matcher.LOCAL_EUVD_SHA256_FILE.write_text(
            hashlib.sha256(self.database_path.read_bytes()).hexdigest() + "\n",
            encoding="ascii",
        )
        matcher.EuvdClient._local_validation_cache = None

    def test_vendor_joint_match_narrows_candidates(self) -> None:
        # Same product_key, two vendors: joint match must return only the
        # record whose vendor matches the SBOM-supplied vendor.
        self._add_second_product("OtherVendor", "EUVD-2024-2")

        async def exercise() -> None:
            wrapper = matcher.EuvdClient()
            async with httpx.AsyncClient() as client:
                items, meta = await wrapper.search(
                    client, "Example Product", "Example"
                )
            self.assertEqual(meta["vendor_filter"], "joint")
            self.assertEqual([item["id"] for item in items], ["EUVD-2024-1"])

        asyncio.run(exercise())

    def test_vendor_fallback_when_joint_match_empty(self) -> None:
        # SBOM vendor disagrees with every EUVD vendor spelling: joint match is
        # empty, so widen back to product-only rather than miss silently
        # (design constraint #5).
        self._add_second_product("OtherVendor", "EUVD-2024-2")

        async def exercise() -> None:
            wrapper = matcher.EuvdClient()
            async with httpx.AsyncClient() as client:
                items, meta = await wrapper.search(
                    client, "Example Product", "UnknownVendor"
                )
            self.assertEqual(meta["vendor_filter"], "fallback")
            self.assertEqual(
                sorted(item["id"] for item in items),
                ["EUVD-2024-1", "EUVD-2024-2"],
            )

        asyncio.run(exercise())

    def test_no_vendor_matches_product_only(self) -> None:
        async def exercise() -> None:
            wrapper = matcher.EuvdClient()
            async with httpx.AsyncClient() as client:
                items, meta = await wrapper.search(client, "Example Product", "")
            self.assertEqual(meta["vendor_filter"], "none")
            self.assertEqual([item["id"] for item in items], ["EUVD-2024-1"])

        asyncio.run(exercise())

    def test_product_candidates_capped_to_top_n(self) -> None:
        # Six product candidates share one product_key with distinct vendors.
        # product_version="*" keeps affected indeterminate so evaluate_item
        # does not drop them via the version gate (L1343); the cap, not
        # version filtering, is what this test exercises.
        for idx in range(2, 7):
            self._add_second_product(f"Vendor{idx}", f"EUVD-2024-{idx}", "*")
        original_cap = matcher.MAX_PRODUCT_CANDIDATES
        matcher.MAX_PRODUCT_CANDIDATES = 3
        try:

            async def exercise() -> None:
                result = await matcher.match_components(
                    [
                        matcher.Component(
                            row_number=2,
                            name="Example Product",
                            version="1.0",
                            vendor="Nomatch",  # no joint match -> fallback -> all 6
                        )
                    ]
                )
                # Cap keeps the top-3 confidence slice. The per-component sort
                # floats 已匹配 + higher CVSS to the front, so capping only
                # drops the low-confidence 需复核 tail — never a confirmed match.
                self.assertEqual(len(result["matches"]), 3)
                self.assertTrue(result["components"][0]["query_truncated"])

            asyncio.run(exercise())
        finally:
            matcher.MAX_PRODUCT_CANDIDATES = original_cap

    def test_hash_mismatch_fails_closed_without_network_query(self) -> None:
        matcher.LOCAL_EUVD_SHA256_FILE.write_text("0" * 64 + "\n", encoding="ascii")
        matcher.EuvdClient._local_validation_cache = None
        status = matcher.EuvdClient.local_snapshot_status()
        self.assertEqual(status["status"], "local_unavailable")
        self.assertEqual(status["snapshot_validation_status"], "hash_mismatch")

        async def exercise() -> None:
            wrapper = matcher.EuvdClient()
            async with httpx.AsyncClient() as client:
                with self.assertRaisesRegex(ValueError, "网络回退未启用"):
                    await wrapper.search(client, "Example Product", "Example")

        asyncio.run(exercise())


if __name__ == "__main__":
    unittest.main()
